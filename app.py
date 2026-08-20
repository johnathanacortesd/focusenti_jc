# ======================================
# Importaciones
# ======================================
import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, NamedStyle, Alignment
from collections import defaultdict, Counter
from difflib import SequenceMatcher
from copy import deepcopy
import datetime
import io
import openai
import re
import time
from unidecode import unidecode
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.cluster import AgglomerativeClustering
import json
import asyncio
import hashlib
from typing import List, Dict, Tuple, Optional, Any
import joblib
import gc
import requests
import os
import zipfile
import xml.etree.ElementTree as ET
import html
from pathlib import Path

# ======================================
# Configuración general
# ======================================
st.set_page_config(
    page_title="Análisis de Sentimiento por Marca",
    page_icon="◈",
    layout="wide",
    initial_sidebar_state="collapsed"
)

OPENAI_MODEL_EMBEDDING     = "text-embedding-3-small"
OPENAI_MODEL_CLASIFICACION = "gpt-4.1-nano-2025-04-14"

CONCURRENT_REQUESTS          = 50
SIMILARITY_THRESHOLD_TONO    = 0.94
SIMILARITY_THRESHOLD_TITULOS = 0.92
MAX_PALABRAS_SUBTEMA         = 5

# ── Umbrales base (corpus grande ≥ 20 noticias) ──────────────────────────────
UMBRAL_SUBTEMA = 0.78
UMBRAL_TEMA    = 0.72
NUM_TEMAS_MAX  = 15

UMBRAL_DEDUP_LABEL           = 0.86
UMBRAL_FUSION_SUBTEMAS       = 0.88
UMBRAL_FUSION_INTERGRUPO     = 0.90
MAX_ITER_FUSION              = 3

UMBRAL_MIN_PERTENENCIA_SUBTEMA = 0.60
UMBRAL_MIN_PERTENENCIA_TEMA    = 0.52

UMBRAL_COHERENCIA_ETIQUETA   = 0.35

MAX_GRUPO_ETIQUETA           = 40

# ── Umbrales mínimos de similitud REAL para agrupar ──────────────────────────
SIM_MINIMA_AGRUPACION_SUBTEMA = 0.90
SIM_MINIMA_KEYWORDS_RARAS     = 0.86   
SIM_MINIMA_FUSION_INTER       = 0.90   

PRICE_INPUT_1M     = 0.10
PRICE_OUTPUT_1M    = 0.40
PRICE_EMBEDDING_1M = 0.02

if 'tokens_input' not in st.session_state: st.session_state['tokens_input']     = 0
if 'tokens_output' not in st.session_state: st.session_state['tokens_output']    = 0
if 'tokens_embedding' not in st.session_state: st.session_state['tokens_embedding'] = 0

STOPWORDS_ES = set("""
a ante bajo cabe con contra de desde durante en entre hacia hasta mediante
para por segun sin so sobre tras y o u e la el los las un una unos unas lo
al del se su sus le les mi mis tu tus nuestro nuestros vuestra vuestras este
esta estos estas ese esa esos esas aquel aquella aquellos aquellas que cual
cuales quien quienes cuyo cuya cuyos cuyas como cuando donde cual es son fue
fueron era eran sera seran seria serian he ha han habia han hay hubo habra
habria estoy esta estan estaba estaban estamos estan estar estare estaria
estuvieron estarian estuvo asi ya mas menos tan tanto cada muy todo toda todos
todas ser haber hacer tener poder deber ir dar ver saber querer llegar pasar
encontrar creer decir poner salir volver seguir llevar sentir cambiar
""".split())

_TRAILING_INCOMPLETE = {
    "de","del","la","el","los","las","un","una","unos","unas","al","su","sus",
    "en","con","sin","por","para","sobre","ante","bajo","contra","desde",
    "entre","hacia","hasta","mediante","tras","y","o","u","e","lo","que","se",
    "como","donde","cuando","cual","cuyo","cuya","cuyos","cuyas",
    "este","esta","estos","estas","ese","esa","esos","esas",
    "aquel","aquella","aquellos","aquellas","cada","todo","toda","todos","todas",
    "otro","otra","otros","otras","nuevo","nueva","nuevos","nuevas",
    "gran","grandes","mayor","mayores","menor","menores","mejor","mejores",
    "peor","peores","primer","primera","segundo","segunda","tercer","tercera",
    "más","mas","muy","tan","tanto","tanta","tantos","tantas",
    "mi","mis","tu","tus","nuestro","nuestra","nuestros","nuestras",
    "a","ha","he","ser","estar","haber","hacer","tener","poder","deber",
    "ir","dar","ver","saber","querer","llegar","pasar","decir","poner",
}

_PATRON_TITULAR = re.compile(
    r"^(nuevo|nueva|anuncia|lanza|presenta|inaugura|llega|abre|inicia|"
    r"logra|alcanza|supera|confirma|destaca|revela|señala|advierte|"
    r"lanzamiento|anuncio|apertura|inicio|presentacion|presentación)\b",
    re.IGNORECASE
)
_PATRON_ESTADO = re.compile(
    r"\b(calma|caos|urgente|hoy|ya|ahora|yesterday|mañana|nuevo|nueva|"
    r"gran|grande|importante|especial|exclusivo)\s*$",
    re.IGNORECASE
)

_TILDE_MAP = {
    "regulacion":"regulación","regulaciones":"regulaciones","innovacion":"innovación",
    "innovaciones":"innovaciones","tecnologia":"tecnología","tecnologias":"tecnologías",
    "tecnologica":"tecnológica","tecnologico":"tecnológico","educacion":"educación",
    "gestion":"gestión","administracion":"administración","informacion":"información",
    "comunicacion":"comunicación","comunicaciones":"comunicaciones","operacion":"operación",
    "operaciones":"operaciones","inversion":"inversión","inversiones":"inversiones",
    "expansion":"expansión","adquisicion":"adquisición","adquisiciones":"adquisiciones",
    "fusion":"fusión","fusiones":"fusiones","transicion":"transición",
    "transformacion":"transformación","digitalizacion":"digitalización",
    "automatizacion":"automatización","modernizacion":"modernización",
    "optimizacion":"optimización","implementacion":"implementación","evaluacion":"evaluación",
    "planificacion":"planificación","organizacion":"organización","atencion":"atención",
    "produccion":"producción","construccion":"construcción","distribucion":"distribución",
    "exportacion":"exportación","importacion":"importación","comercializacion":"comercialización",
    "negociacion":"negociación","negociaciones":"negociaciones","participacion":"participación",
    "colaboracion":"colaboración","asociacion":"asociación","integracion":"integración",
    "relacion":"relación","relaciones":"relaciones","situacion":"situación",
    "condicion":"condición","condiciones":"condiciones","solucion":"solución",
    "soluciones":"soluciones","prevencion":"prevención","proteccion":"protección",
    "fiscalizacion":"fiscalización","sancion":"sanción","sanciones":"sanciones",
    "investigacion":"investigación","investigaciones":"investigaciones","accion":"acción",
    "acciones":"acciones","direccion":"dirección","decision":"decisión",
    "decisiones":"decisiones","eleccion":"elección","elecciones":"elecciones",
    "votacion":"votación","aprobacion":"aprobación","legislacion":"legislación",
    "reclamacion":"reclamación","reclamaciones":"reclamaciones","obligacion":"obligación",
    "obligaciones":"obligaciones","inflacion":"inflación","tributacion":"tributación",
    "financiera":"financiera","financiero":"financiero","economica":"económica",
    "economico":"económico","economia":"economía","credito":"crédito",
    "creditos":"créditos","prestamo":"préstamo","prestamos":"préstamos",
    "interes":"interés","comision":"comisión","comisiones":"comisiones",
    "politica":"política","politicas":"políticas","politico":"político",
    "publica":"pública","publico":"público","estrategia":"estrategia",
    "estrategica":"estratégica","estrategico":"estratégico","logistica":"logística",
    "analisis":"análisis","diagnostico":"diagnóstico","indice":"índice",
    "vehiculo":"vehículo","vehiculos":"vehículos","electrico":"eléctrico",
    "electrica":"eléctrica","energia":"energía","energetica":"energética",
    "petroleo":"petróleo","mineria":"minería","agricola":"agrícola",
    "biologica":"biológica","ecologica":"ecológica","inclusion":"inclusión",
    "exclusion":"exclusión","pension":"pensión","pensiones":"pensiones",
    "jubilacion":"jubilación","compensacion":"compensación","remuneracion":"remuneración",
    "contratacion":"contratación","capacitacion":"capacitación","formacion":"formación",
    "certificacion":"certificación","habilitacion":"habilitación","autorizacion":"autorización",
    "concesion":"concesión","licitacion":"licitación","migracion":"migración",
    "poblacion":"población","recaudacion":"recaudación","asignacion":"asignación",
    "corporacion":"corporación","fundacion":"fundación","institucion":"institución",
    "instituciones":"instituciones","region":"región","unico":"único","unica":"única",
    "ultimo":"último","ultima":"última","proximo":"próximo","basico":"básico",
    "basica":"básica","historico":"histórico","historica":"histórica",
    "medico":"médico","medica":"médica","farmaceutica":"farmacéutica",
    "clinica":"clínica","numero":"número","telefono":"teléfono","telefonia":"telefonía",
    "movil":"móvil","moviles":"móviles","codigo":"código","informatica":"informática",
    "electronica":"electrónica","robotica":"robótica","ciberseguridad":"ciberseguridad",
    "trafico":"tráfico","transito":"tránsito","aereo":"aéreo","maritimo":"marítimo",
    "turistica":"turística","turistico":"turístico","gastronomia":"gastrónomía",
    "academica":"académica","academico":"académico","pedagogica":"pedagógica",
    "cientifica":"científica","cientifico":"científico","juridica":"jurídica",
    "juridico":"jurídico","constitucion":"constitución","resolucion":"resolución",
    "notificacion":"notificación","programacion":"programación","actualizacion":"actualización",
    "verificacion":"verificación","validacion":"validación","liquidacion":"liquidación",
    "facturacion":"facturación","evasion":"evasión","corrupcion":"corrupción",
    "deforestacion":"deforestación","contaminacion":"contaminación","conservacion":"conservación",
    "restauracion":"restauración","rehabilitacion":"rehabilitación","renovacion":"renovación",
    "ampliacion":"ampliación","inauguracion":"inauguración","celebracion":"celebración",
    "clasificacion":"clasificación","eliminacion":"eliminación","motivacion":"motivación",
    "satisfaccion":"satisfacción","reputacion":"reputación","disposicion":"disposición",
}

_ENIE_MAP = {
    "desempeno":"desempeño","desempenos":"desempeños","empeno":"empeño","empenos":"empeños",
    "ensenanza":"enseñanza","ensenanzas":"enseñanzas","diseno":"diseño","disenos":"diseños",
    "disenador":"diseñador","disenadora":"diseñadora","disenadores":"diseñadores",
    "nino":"niño","nina":"niña","ninos":"niños","ninas":"niñas","ninez":"niñez",
    "ano":"año","anos":"años","danio":"daño","danios":"daños","dano":"daño","danos":"daños",
    "danino":"dañino","danina":"dañina","montana":"montaña","montanas":"montañas",
    "espana":"España","espanol":"español","espanola":"española","espanoles":"españoles",
    "companero":"compañero","companera":"compañera","companeros":"compañeros","companeras":"compañeras",
    "compania":"compañía","companias":"compañías","acompanamiento":"acompanamiento",
    "cana":"caña","canas":"cañas","banio":"baño","banios":"baños","bano":"baño","banos":"baños",
    "pena":"peña","penas":"peñas","penon":"peñón","senor":"señor","senora":"señora",
    "senores":"señores","senoras":"señoras","senal":"señal","senales":"señales",
    "senalizacion":"señalización","pequeno":"pequeño","pequena":"pequeña",
    "pequenos":"pequeños","pequenas":"peñas","sueno":"sueño","suenos":"sueños",
    "dueno":"dueño","duena":"dueña","duenos":"dueños","duenas":"dueñas",
    "otono":"otoño","punio":"puño","punios":"puños","puno":"puño",
    "canon":"cañón","canones":"cañones","manana":"mañana","mananas":"mañanas",
    "cabana":"cabaña","cabanas":"cabañas","banera":"bañera","vinedo":"viñedo",
    "vinedos":"viñedos","rebano":"rebaño","rebanos":"rebaños","extrano":"extraño",
    "extrana":"extraña","extranos":"extraños","extranas":"extrañas",
    "enganio":"engaño","engano":"engaño","enganos":"engaños","tamanio":"tamaño",
    "tamano":"tamaño","tamanos":"tamaños","muneca":"muñeca","munecas":"muñecas",
    "cunado":"cuñado","cunada":"cuñada","cunados":"cuñados","albanil":"albañil",
    "albaniles":"albañiles","narino":"Nariño","quindio":"Quindío",
    "ibanez":"Ibáñez","nunez":"Núñez","munoz":"Muñoz","ordonez":"Ordóñez",
    "yanez":"Yáñez","castaneda":"Castañeda","penalosa":"Peñalosa",
    "vineta":"viñeta","vinetas":"viñetas","banado":"bañado","banada":"bañada",
    "rinon":"riñón","rinones":"riñones","panial":"pañal","paniales":"pañales",
    "panal":"pañal","panales":"pañales","arana":"araña","aranas":"arañas",
    "pestana":"pestaña","pestanas":"pestañas","guino":"guiño","guinos":"guiños",
    "munequera":"muñequera","lenador":"leñador","lenadores":"leñadores",
    "resena":"reseña","resenas":"reseñas","panuelo":"pañuelo","panuelos":"pañuelos",
    "companerismo":"compañerismo","desengano":"desengaño","lenio":"leño","leno":"leño",
}

def corregir_tildes(texto: str) -> str:
    if not texto: return texto
    palabras = texto.split()
    resultado = []
    for p in palabras:
        low = p.lower()
        if low in _TILDE_MAP:
            c = _TILDE_MAP[low]
            if p[0].isupper() and not c[0].isupper(): c = c[0].upper() + c[1:]
            resultado.append(c)
        elif low in _ENIE_MAP:
            c = _ENIE_MAP[low]
            if p[0].isupper() and not c[0].isupper(): c = c[0].upper() + c[1:]
            resultado.append(c)
        else:
            resultado.append(p)
    return " ".join(resultado)


# ======================================
# CSS
# ======================================
def load_custom_css():
    st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Google+Sans:wght@400;500;700&family=Google+Sans+Text:wght@400;500;700&family=Roboto+Mono:wght@400;500&display=swap');
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
:root {
    --bg:#f8f9fa;--s1:#ffffff;--s2:#f1f3f4;--s3:#e8eaed;
    --border:#dadce0;--border2:#bdc1c6;--border-focus:#f97316;
    --text:#202124;--text2:#3c4043;--text3:#5f6368;--text4:#9aa0a6;
    --accent:#f97316;--accent2:#ea580c;--accent3:#c2410c;
    --accent-bg:#fff7ed;--accent-bg2:#ffedd5;--accent-bdr:#fed7aa;
    --green:#059669;--green2:#047857;--green-bg:#ecfdf5;--green-bdr:#a7f3d0;
    --red:#dc2626;--amber:#d97706;--blue:#1a73e8;
    --r:8px;--r2:12px;--r3:16px;--r4:20px;
    --shadow-sm:0 1px 2px rgba(60,64,67,0.1),0 1px 3px rgba(60,64,67,0.08);
    --shadow-md:0 1px 3px rgba(60,64,67,0.12),0 4px 8px rgba(60,64,67,0.08);
    --shadow-lg:0 2px 6px rgba(60,64,67,0.1),0 8px 24px rgba(60,64,67,0.1);
    --transition:all 0.2s cubic-bezier(0.4,0,0.2,1);
}
html,body,[data-testid="stApp"]{
    background:var(--bg)!important;color:var(--text)!important;
    font-family:'Google Sans Text','Inter',-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;
    font-size:14px;-webkit-font-smoothing:antialiased;letter-spacing:0.01em;
}
#MainMenu,footer,header{visibility:hidden}.stDeployButton{display:none}
.block-container{padding-top:1rem!important;padding-bottom:0!important}
[data-testid="stAppViewBlockContainer"]{padding-top:1rem!important}
.app-header{background:var(--s1);border:1px solid var(--border);border-radius:var(--r3);padding:1rem 1.5rem;margin-bottom:1rem;display:flex;align-items:center;gap:1rem;box-shadow:var(--shadow-sm);position:relative;overflow:hidden;}
.app-header-icon{width:40px;height:40px;background:linear-gradient(135deg,#f97316,#ea580c);border-radius:12px;display:flex;align-items:center;justify-content:center;font-size:1.2rem;color:white;flex-shrink:0;box-shadow:0 2px 8px rgba(249,115,22,0.3);}
.app-header-text{flex:1}
.app-header-title{font-family:'Google Sans',sans-serif;font-size:1.25rem;font-weight:700;color:var(--text);letter-spacing:-0.01em;line-height:1.3}
.app-header-version{font-family:'Roboto Mono',monospace;font-size:0.65rem;color:var(--text3);letter-spacing:0.03em;margin-top:0.15rem}
.app-header-badge{background:var(--accent-bg);border:1px solid var(--accent-bdr);color:var(--accent2);font-family:'Roboto Mono',monospace;font-size:0.6rem;font-weight:500;padding:0.25rem 0.75rem;border-radius:100px;letter-spacing:0.04em;text-transform:uppercase;white-space:nowrap;}
[data-testid="stTabs"] [data-testid="stTabsList"]{background:var(--s1)!important;border:1px solid var(--border)!important;border-radius:var(--r2)!important;padding:4px!important;gap:4px!important;box-shadow:var(--shadow-sm)!important;margin-bottom:0.75rem!important;}
[data-testid="stTabs"] button[data-baseweb="tab"]{font-family:'Google Sans',sans-serif!important;font-size:0.88rem!important;font-weight:500!important;color:var(--text2)!important;border-radius:var(--r)!important;padding:0.45rem 1.2rem!important;border:none!important;background:transparent!important;transition:var(--transition)!important;}
[data-testid="stTabs"] button[data-baseweb="tab"]:hover{background:var(--s2)!important;color:var(--text)!important}
[data-testid="stTabs"] button[data-baseweb="tab"][aria-selected="true"]{background:var(--accent-bg)!important;color:var(--accent2)!important;border:1px solid var(--accent-bdr)!important;font-weight:700!important;}
.metrics-grid{display:grid;grid-template-columns:repeat(5,1fr);gap:0.6rem;margin:0.8rem 0}
.metric-card{background:var(--s1);border:1px solid var(--border);border-radius:var(--r2);padding:0.8rem 0.6rem;text-align:center;transition:var(--transition);box-shadow:var(--shadow-sm);position:relative;overflow:hidden;}
.metric-card::before{content:'';position:absolute;top:0;left:0;right:0;height:3px;border-radius:var(--r2) var(--r2) 0 0}
.metric-card.m-total::before{background:linear-gradient(90deg,#5f6368,#9aa0a6)}
.metric-card.m-unique::before{background:linear-gradient(90deg,#059669,#34d399)}
.metric-card.m-dup::before{background:linear-gradient(90deg,#f59e0b,#fbbf24)}
.metric-card.m-time::before{background:linear-gradient(90deg,#1a73e8,#4285f4)}
.metric-card.m-cost::before{background:linear-gradient(90deg,#f97316,#fb923c)}
.metric-card:hover{transform:translateY(-2px);box-shadow:var(--shadow-lg)}
.metric-val{font-family:'Google Sans',sans-serif;font-size:1.5rem;font-weight:700;line-height:1;margin-bottom:0.3rem;letter-spacing:-0.01em}
.metric-lbl{font-family:'Roboto Mono',monospace;font-size:0.62rem;color:var(--text3);text-transform:uppercase;letter-spacing:0.08em;font-weight:500}
[data-testid="stForm"]{background:var(--s1)!important;border:1px solid var(--border)!important;border-radius:var(--r3)!important;padding:1.2rem 1.5rem!important;box-shadow:var(--shadow-md)!important;}
.sec-label{font-family:'Google Sans',sans-serif;font-size:0.72rem;font-weight:700;color:var(--text2);letter-spacing:0.08em;text-transform:uppercase;padding-bottom:0.3rem;border-bottom:2px solid var(--s3);margin:0.8rem 0 0.5rem;display:flex;align-items:center;gap:0.5rem;}
.sec-label::before{content:'';display:inline-block;width:3px;height:12px;background:linear-gradient(180deg,#f97316,#ea580c);border-radius:2px}
.upload-zone{display:grid;grid-template-columns:repeat(3,1fr);gap:0.6rem;margin:0.3rem 0}
.upload-zone-card{background:var(--s1);border:1.5px dashed var(--border);border-radius:var(--r2);padding:0.6rem 0.8rem;display:flex;align-items:center;gap:0.6rem;transition:var(--transition);}
.upload-zone-card:hover{border-color:var(--accent);border-style:solid;transform:translateY(-1px);box-shadow:var(--shadow-md)}
.upload-zone-icon{width:32px;height:32px;border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:1rem;flex-shrink:0;}
.upload-zone-icon.uz-dossier{background:#fff7ed;color:#f97316}
.upload-zone-icon.uz-region{background:#ecfdf5;color:#059669}
.upload-zone-icon.uz-internet{background:#eff6ff;color:#1a73e8}
.upload-zone-text{flex:1;min-width:0}
.upload-zone-title{font-family:'Google Sans',sans-serif;font-size:0.82rem;font-weight:700;color:var(--text);line-height:1.2}
.upload-zone-desc{font-size:0.7rem;color:var(--text3);line-height:1.3}
[data-testid="stFileUploader"]{background:var(--s1)!important;border:1.5px dashed var(--border)!important;border-radius:var(--r)!important;padding:0.4rem 0.6rem!important;transition:var(--transition)!important;min-height:auto!important;}
[data-testid="stFileUploader"]:hover{border-color:var(--accent)!important;border-style:solid!important;background:var(--accent-bg)!important;}
[data-testid="stFileUploader"] section{padding:0.2rem!important}
[data-testid="stFileUploader"] section>div{font-size:0.78rem!important;color:var(--text2)!important}
[data-testid="stFileUploader"] section small{font-size:0.7rem!important;color:var(--text3)!important}
[data-testid="stFileUploader"] button{background:var(--accent-bg)!important;border:1px solid var(--accent-bdr)!important;color:var(--accent2)!important;font-weight:500!important;font-size:0.75rem!important;border-radius:100px!important;padding:0.25rem 0.8rem!important;font-family:'Google Sans',sans-serif!important;transition:var(--transition)!important;}
[data-testid="stFileUploader"] button:hover{background:var(--accent)!important;color:white!important;border-color:var(--accent)!important}
[data-testid="stTextInput"] input,[data-testid="stTextArea"] textarea{background:var(--s1)!important;border:1.5px solid var(--border)!important;color:var(--text)!important;border-radius:var(--r)!important;font-family:'Google Sans Text',sans-serif!important;font-size:0.9rem!important;padding:0.5rem 0.75rem!important;transition:var(--transition)!important;}
[data-testid="stTextInput"] input:focus,[data-testid="stTextArea"] textarea:focus{border-color:var(--accent)!important;box-shadow:0 0 0 3px rgba(249,115,22,0.12)!important;}
[data-testid="stTextInput"] input::placeholder,[data-testid="stTextArea"] textarea::placeholder{color:var(--text4)!important;font-size:0.85rem!important;}
label[data-testid="stWidgetLabel"] p{font-family:'Google Sans',sans-serif!important;color:var(--text2)!important;font-size:0.82rem!important;font-weight:500!important;margin-bottom:0.15rem!important;}
.stButton>button,[data-testid="stDownloadButton"]>button{background:var(--s1)!important;border:1.5px solid var(--border)!important;color:var(--text)!important;border-radius:100px!important;font-family:'Google Sans',sans-serif!important;font-weight:500!important;font-size:0.88rem!important;transition:var(--transition)!important;padding:0.5rem 1.2rem!important;box-shadow:none!important;}
.stButton>button:hover,[data-testid="stDownloadButton"]>button:hover{border-color:var(--accent)!important;color:var(--accent2)!important;background:var(--accent-bg)!important;box-shadow:var(--shadow-sm)!important;transform:translateY(-1px)!important;}
.stButton>button[kind="primary"],[data-testid="stDownloadButton"]>button[kind="primary"]{background:var(--accent)!important;border:none!important;color:#fff!important;font-weight:500!important;font-size:0.92rem!important;padding:0.6rem 1.5rem!important;box-shadow:0 1px 3px rgba(249,115,22,0.3),0 4px 12px rgba(249,115,22,0.15)!important;letter-spacing:0.01em!important;}
.stButton>button[kind="primary"]:hover,[data-testid="stDownloadButton"]>button[kind="primary"]:hover{background:var(--accent2)!important;box-shadow:0 2px 6px rgba(234,88,12,0.35),0 8px 24px rgba(234,88,12,0.18)!important;transform:translateY(-1px)!important;color:#fff!important;}
[data-testid="stRadio"] label{font-family:'Google Sans Text',sans-serif!important;color:var(--text)!important;font-size:0.88rem!important;font-weight:400!important;}
[data-testid="stRadio"]{margin-bottom:0!important}
[data-testid="stRadio"]>div{gap:0!important}
[data-testid="stStatus"]{background:var(--s1)!important;border:1px solid var(--border)!important;border-radius:var(--r2)!important;font-family:'Roboto Mono',monospace!important;font-size:0.8rem!important;}
[data-testid="stAlert"]{background:var(--s1)!important;border:1px solid var(--border)!important;border-radius:var(--r2)!important;color:var(--text2)!important;font-size:0.85rem!important;padding:0.6rem 0.8rem!important;}
.success-banner{background:linear-gradient(135deg,#ecfdf5,#d1fae5);border:1px solid var(--green-bdr);border-left:4px solid var(--green);border-radius:var(--r2);padding:0.8rem 1.2rem;margin:0.5rem 0 0.8rem;display:flex;align-items:center;gap:0.8rem;}
.success-icon{width:34px;height:34px;background:linear-gradient(135deg,#059669,#047857);border-radius:50%;display:flex;align-items:center;justify-content:center;color:white;font-size:1rem;flex-shrink:0;}
.success-title{font-family:'Google Sans',sans-serif;font-size:1rem;font-weight:700;color:#047857;margin-bottom:0.1rem}
.success-sub{font-size:0.8rem;color:var(--text2)}
.auth-wrap{max-width:380px;margin:8vh auto 0;text-align:center}
.auth-icon{width:60px;height:60px;background:linear-gradient(135deg,#f97316,#ea580c);border-radius:16px;display:inline-flex;align-items:center;justify-content:center;font-size:1.6rem;color:white;margin-bottom:1rem;box-shadow:0 4px 16px rgba(249,115,22,0.3);}
.auth-title{font-family:'Google Sans',sans-serif;font-size:1.5rem;font-weight:700;color:var(--text);margin-bottom:0.3rem}
.auth-sub{font-size:0.85rem;color:var(--text3);margin-bottom:2rem}
.cluster-info{background:var(--accent-bg);border:1px solid var(--accent-bdr);border-radius:var(--r);padding:0.5rem 0.8rem;margin:0.4rem 0;font-family:'Roboto Mono',monospace;font-size:0.68rem;color:var(--text2);line-height:1.6;}
.cluster-info b{color:var(--accent2);font-size:0.72rem}
.config-badge{display:inline-flex;align-items:center;gap:0.4rem;background:var(--s2);border:1px solid var(--border);border-radius:100px;padding:0.2rem 0.7rem;font-family:'Roboto Mono',monospace;font-size:0.62rem;color:var(--text3);margin-bottom:0.6rem;}
[data-testid="stProgressBar"]>div>div{background:linear-gradient(90deg,#f97316,#fb923c,#fdba74)!important;border-radius:100px!important;height:5px!important;}
[data-testid="stDataFrame"]{border:1px solid var(--border)!important;border-radius:var(--r2)!important;box-shadow:var(--shadow-sm)!important;overflow:hidden!important;}
::-webkit-scrollbar{width:6px;height:6px}
::-webkit-scrollbar-track{background:var(--s2);border-radius:3px}
::-webkit-scrollbar-thumb{background:var(--border2);border-radius:3px}
::-webkit-scrollbar-thumb:hover{background:var(--accent)}
.footer{font-family:'Roboto Mono',monospace;font-size:0.6rem;color:var(--text4);text-align:center;padding:0.8rem 0 0.5rem;letter-spacing:0.04em;border-top:1px solid var(--s3);margin-top:1rem;}
.stElementContainer{margin-bottom:0!important}
[data-testid="stVerticalBlock"]>div{gap:0.3rem!important}
[data-testid="stHorizontalBlock"]>div{gap:0.4rem!important}
hr{border-color:var(--s3)!important;margin:0.5rem 0!important}
[data-testid="stSelectbox"]>div>div{font-family:'Google Sans Text',sans-serif!important;font-size:0.88rem!important;color:var(--text)!important;}
@media(max-width:768px){
    .metrics-grid{grid-template-columns:repeat(2,1fr)}
    .upload-zone{grid-template-columns:1fr}
    .app-header{flex-direction:column;text-align:center;gap:0.5rem;padding:1rem}
}
</style>
""", unsafe_allow_html=True)


# ======================================
# Umbrales adaptativos según tamaño del corpus
# ======================================
def _umbrales_adaptativos(n: int) -> dict:
    if n <= 5:
        return dict(
            subtema=0.93,
            tema=0.85,
            dedup_label=0.90,
            fusion_subtemas=0.92,
            fusion_intergrupo=0.95,
            min_pertenencia_subtema=0.80,
            min_pertenencia_tema=0.75,
            coherencia_etiqueta=0.50,
            sim_minima_agrupacion=0.93,
            sim_minima_keywords=0.93,
            max_iter_fusion=1,
            num_temas_max=n,
            usar_paso2b=False,
            usar_fusion_iterativa=False,
        )
    elif n <= 10:
        return dict(
            subtema=0.90,
            tema=0.84,
            dedup_label=0.88,
            fusion_subtemas=0.90,
            fusion_intergrupo=0.93,
            min_pertenencia_subtema=0.72,
            min_pertenencia_tema=0.65,
            coherencia_etiqueta=0.42,
            sim_minima_agrupacion=0.90,
            sim_minima_keywords=0.90,
            max_iter_fusion=2,
            num_temas_max=min(n, 5),
            usar_paso2b=False,
            usar_fusion_iterativa=False,
        )
    elif n <= 20:
        return dict(
            subtema=0.87,
            tema=0.82,
            dedup_label=0.86,
            fusion_subtemas=0.88,
            fusion_intergrupo=0.91,
            min_pertenencia_subtema=0.66,
            min_pertenencia_tema=0.58,
            coherencia_etiqueta=0.38,
            sim_minima_agrupacion=0.87,
            sim_minima_keywords=0.87,
            max_iter_fusion=3,
            num_temas_max=min(n // 2, NUM_TEMAS_MAX),
            usar_paso2b=True,
            usar_fusion_iterativa=True,
        )
    else:
        return dict(
            subtema=UMBRAL_SUBTEMA,
            tema=UMBRAL_TEMA,
            dedup_label=UMBRAL_DEDUP_LABEL,
            fusion_subtemas=UMBRAL_FUSION_SUBTEMAS,
            fusion_intergrupo=UMBRAL_FUSION_INTERGRUPO,
            min_pertenencia_subtema=UMBRAL_MIN_PERTENENCIA_SUBTEMA,
            min_pertenencia_tema=UMBRAL_MIN_PERTENENCIA_TEMA,
            coherencia_etiqueta=UMBRAL_COHERENCIA_ETIQUETA,
            sim_minima_agrupacion=SIM_MINIMA_AGRUPACION_SUBTEMA,
            sim_minima_keywords=SIM_MINIMA_KEYWORDS_RARAS,
            max_iter_fusion=MAX_ITER_FUSION,
            num_temas_max=NUM_TEMAS_MAX,
            usar_paso2b=True,
            usar_fusion_iterativa=True,
        )


# ======================================
# Caché Global de Embeddings
# ======================================
class EmbeddingCache:
    def __init__(self):
        self._cache: Dict[str, List[float]] = {}
        self._hits = 0
        self._misses = 0

    def _key(self, text):
        return hashlib.md5(text[:2000].encode('utf-8', errors='ignore')).hexdigest()

    def get(self, text):
        k = self._key(text)
        if k in self._cache:
            self._hits += 1
            return self._cache[k]
        self._misses += 1
        return None

    def put(self, text, emb):
        self._cache[self._key(text)] = emb

    def get_many(self, textos):
        results = [None] * len(textos)
        missing = []
        for i, t in enumerate(textos):
            c = self.get(t)
            if c is not None:
                results[i] = c
            else:
                missing.append(i)
        return results, missing

    def stats(self):
        total = self._hits + self._misses
        rate = (self._hits / total * 100) if total > 0 else 0
        return f"Cache: {self._hits} hits, {self._misses} misses ({rate:.0f}%)"

    def clear(self):
        self._cache.clear()
        self._hits = 0
        self._misses = 0

if '_emb_cache' not in st.session_state:
    st.session_state['_emb_cache'] = EmbeddingCache()

def get_embedding_cache():
    return st.session_state['_emb_cache']

# ======================================
# Configuración vía Google Sheets (CSV público)
# ======================================
CONFIG_CACHE_TTL = 300  # segundos

@st.cache_data(ttl=CONFIG_CACHE_TTL, show_spinner=False)
def _fetch_map_from_csv(csv_url: str) -> dict:
    df = pd.read_csv(csv_url, header=None, dtype=str)
    df = df.dropna(how="all")
    mapping = pd.Series(
        df.iloc[:, 1].values,
        index=df.iloc[:, 0].astype(str).str.lower().str.strip()
    ).to_dict()
    mapping = {k: v for k, v in mapping.items() if k not in ("nan", "")}
    return mapping

def load_config_from_sheets():
    regiones_url = st.secrets.get("REGIONES_CSV_URL")
    internet_url = st.secrets.get("INTERNET_CSV_URL")

    if not regiones_url or not internet_url:
        st.error(
            "❌ Faltan las URLs de configuración. Agrega REGIONES_CSV_URL e "
            "INTERNET_CSV_URL en los Secrets de la app."
        )
        st.stop()

    try:
        region_map = _fetch_map_from_csv(regiones_url)
        internet_map = _fetch_map_from_csv(internet_url)
    except Exception as e:
        st.error(f"❌ No se pudo leer la configuración desde Google Sheets: {e}")
        st.stop()

    return region_map, internet_map

def refresh_config_cache():
    _fetch_map_from_csv.clear()


# ======================================
# Funciones Auxiliares de Limpieza, Enlaces y Conversión
# ======================================

def check_password():
    if st.session_state.get("password_correct", False):
        return True
    st.markdown("""
    <div class="auth-wrap">
        <div class="auth-icon">◈</div>
        <div class="auth-title">Sistema de Análisis</div>
        <div class="auth-sub">Ingresa tus credenciales para continuar</div>
    </div>""", unsafe_allow_html=True)
    _, col, _ = st.columns([1, 2, 1])
    with col:
        with st.form("pw"):
            pw = st.text_input("Contraseña", type="password", placeholder="Ingresa tu contraseña")
            if st.form_submit_button("Ingresar", use_container_width=True, type="primary"):
                if pw == st.secrets.get("APP_PASSWORD", "INVALID"):
                    st.session_state["password_correct"] = True
                    st.rerun()
                else:
                    st.error("Contraseña incorrecta")
    return False

def call_with_retries(fn, *a, **kw):
    d = 1
    for att in range(3):
        try:
            return fn(*a, **kw)
        except Exception as e:
            if att == 2: raise e
            time.sleep(d)
            d *= 2

async def acall_with_retries(fn, *a, **kw):
    d = 1
    for att in range(3):
        try:
            return await fn(*a, **kw)
        except Exception as e:
            if att == 2: raise e
            await asyncio.sleep(d)
            d *= 2

def norm_key(text):
    if text is None: return ""
    return re.sub(r"[^a-z0-9]+", "", unidecode(str(text).strip().lower()))

def capitalizar_etiqueta(tema):
    if not tema or not tema.strip(): return "Sin tema"
    tema = tema.strip().lower()
    tema = corregir_tildes(tema)
    return tema[0].upper() + tema[1:]

def _frase_esta_completa(texto):
    if not texto or not texto.strip(): return False
    palabras = texto.strip().split()
    if not palabras: return False
    ultima = palabras[-1].lower().rstrip(".,;:!?")
    return unidecode(ultima) not in _TRAILING_INCOMPLETE and len(ultima) > 1

def _recortar_frase_completa(texto, max_palabras=7):
    if not texto: return "Sin tema"
    palabras = texto.strip().split()
    if len(palabras) > max_palabras: palabras = palabras[:max_palabras]
    while palabras and unidecode(palabras[-1].lower().rstrip(".,;:!?")) in _TRAILING_INCOMPLETE:
        palabras.pop()
    if not palabras: return texto.strip().split()[0] if texto.strip() else "Sin tema"
    return " ".join(palabras)

def limpiar_tema(tema):
    if not tema: return "Sin tema"
    tema = tema.strip().strip('"\'')
    for px in ["subtema:", "tema:", "categoría:", "categoria:", "category:"]:
        if tema.lower().startswith(px): tema = tema[len(px):].strip()
    tema = _recortar_frase_completa(tema, max_palabras=MAX_PALABRAS_SUBTEMA)
    return capitalizar_etiqueta(tema) if tema else "Sin tema"

def limpiar_tema_geografico(tema, marca, aliases):
    if not tema: return "Sin tema"
    tl = unidecode(tema.lower())
    for n in [marca] + [a for a in aliases if a]:
        patron = r'\b' + re.escape(unidecode(n.strip().lower())) + r'\b'
        tl = re.sub(patron, '', tl)
    frases_eliminar = [
        "en colombia", "de colombia", "del pais", "en el pais",
        "territorio nacional", "a nivel nacional", "en todo el pais",
    ]
    for frase in frases_eliminar:
        tl = re.sub(r'\b' + re.escape(frase) + r'\b', '', tl)
    tl = re.sub(r'\s+', ' ', tl).strip()
    if not tl: return "Sin tema"
    tokens_orig = tema.split()
    tokens_norm = unidecode(tema.lower()).split()
    norm_disponibles = tl.split()
    resultado_tokens = []
    for orig, norm in zip(tokens_orig, tokens_norm):
        if norm_disponibles and norm == norm_disponibles[0]:
            resultado_tokens.append(orig)
            norm_disponibles.pop(0)
    resultado = " ".join(resultado_tokens).strip()
    resultado = corregir_tildes(resultado) if resultado else ""
    return limpiar_tema(resultado) if resultado.strip() else "Sin tema"

def string_norm_label(s):
    if not s: return ""
    s = unidecode(s.lower())
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    return " ".join(t for t in s.split() if t not in STOPWORDS_ES)

_ACCIONES_OPUESTAS = [
    ({"aprobacion", "aprueba", "apoyo", "acuerdo", "aval", "respaldo"}, {"rechazo", "rechaza", "desacuerdo", "oposicion", "critica"}),
    ({"aumento", "crecimiento", "alza", "subida", "incremento"}, {"caida", "reduccion", "baja", "disminucion", "descenso"}),
    ({"apertura", "inauguracion", "inicio", "lanzamiento", "estreno"}, {"cierre", "suspension", "fin", "clausura", "cancelacion"}),
    ({"exito", "logro", "triunfo", "premio", "reconocimiento"}, {"fracaso", "derrota", "problema", "crisis", "sancion"}),
    ({"demanda", "denuncia", "investigacion", "sancion", "multa"}, {"absolucion", "archivo", "exoneracion", "acuerdo"}),
]

_TOKENS_DEBILES_AGRUPACION = STOPWORDS_ES | {
    "noticia", "noticias", "informe", "informacion", "comunicado", "anuncio",
    "colombia", "pais", "nacional", "regional", "local", "sector", "sectores",
    "empresa", "empresas", "entidad", "entidades", "autoridad", "autoridades",
    "gobierno", "alcaldia", "gobernacion", "ministerio", "nuevo", "nueva",
    "nuevos", "nuevas", "plan", "programa", "proyecto", "iniciativa",
    "actividad", "actividades", "gestion", "tema", "caso", "casos",
}

def _tokens_distintivos(texto: str, min_len: int = 4) -> set:
    norm = string_norm_label(texto)
    return {
        t for t in norm.split()
        if len(t) >= min_len and t not in _TOKENS_DEBILES_AGRUPACION and not t.isdigit()
    }

def _overlap_distintivo(a: str, b: str) -> float:
    ta, tb = _tokens_distintivos(a), _tokens_distintivos(b)
    if not ta or not tb: return 0.0
    return len(ta & tb) / max(1, min(len(ta), len(tb)))

def _hay_conflicto_accion(a: str, b: str) -> bool:
    ta, tb = _tokens_distintivos(a, min_len=3), _tokens_distintivos(b, min_len=3)
    for grupo_a, grupo_b in _ACCIONES_OPUESTAS:
        if (ta & grupo_a and tb & grupo_b) or (ta & grupo_b and tb & grupo_a):
            return True
    return False

def _etiquetas_compatibles(a: str, b: str, min_overlap: float = 0.45) -> bool:
    na, nb = string_norm_label(a), string_norm_label(b)
    if not na or not nb: return False
    if _hay_conflicto_accion(na, nb): return False
    if SequenceMatcher(None, na, nb).ratio() >= 0.90: return True
    return _overlap_distintivo(na, nb) >= min_overlap

def _grupos_contenido_compatibles(
    textos_a: list,
    textos_b: list,
    etiqueta_a: str = "",
    etiqueta_b: str = "",
    min_sim: float = 0.88,
    min_overlap: float = 0.20,
) -> bool:
    muestra_a = [str(t) for t in textos_a[:20] if str(t).strip()]
    muestra_b = [str(t) for t in textos_b[:20] if str(t).strip()]
    if not muestra_a or not muestra_b: return False
    texto_a = " ".join(muestra_a)[:2500]
    texto_b = " ".join(muestra_b)[:2500]
    if _hay_conflicto_accion(f"{etiqueta_a} {texto_a}", f"{etiqueta_b} {texto_b}"):
        return False
    overlap = _overlap_distintivo(f"{etiqueta_a} {texto_a}", f"{etiqueta_b} {texto_b}")
    labels_muy_cercanas = _etiquetas_compatibles(etiqueta_a, etiqueta_b, min_overlap=0.55)
    if overlap < min_overlap and not labels_muy_cercanas:
        return False
    embs = get_embeddings_batch([texto_a, texto_b])
    if len(embs) < 2 or embs[0] is None or embs[1] is None:
        return labels_muy_cercanas and overlap >= min_overlap
    sim = cosine_similarity(
        np.array(embs[0]).reshape(1, -1),
        np.array(embs[1]).reshape(1, -1)
    )[0][0]
    return sim >= min_sim

def _validar_estructura_subtema(etiqueta: str) -> bool:
    if not etiqueta or len(etiqueta.split()) < 2: return False
    if len(etiqueta.split()) > MAX_PALABRAS_SUBTEMA: return False
    if _PATRON_TITULAR.match(etiqueta): return False
    if _PATRON_ESTADO.search(etiqueta): return False
    palabras = etiqueta.split()
    if len(palabras) <= 4:
        nexos = {
            "de","del","para","sobre","en","con","por","ante","hacia",
            "entre","sin","al","las","los","una","uno","que","como",
            "y","o","a","e","u",
        }
        tiene_nexo = any(unidecode(p.lower().rstrip(".,;:!?")) in nexos for p in palabras[1:])
        if not tiene_nexo: return False
    return True

def _es_nombre_o_fragmento_marca(etiqueta: str, marca: str, aliases=None) -> bool:
    """Detecta etiquetas que solo repiten total o parcialmente el nombre de la marca."""
    vacias = {"de", "del", "la", "el", "los", "las", "y", "e", "grupo"}
    tokens_etiqueta = {t for t in _normalizar_mencion(etiqueta).split() if t not in vacias}
    if not tokens_etiqueta:
        return True
    for nombre in _variantes_marca(marca, aliases):
        tokens_marca = {t for t in _normalizar_mencion(nombre).split() if t not in vacias}
        if not tokens_marca:
            continue
        comunes = tokens_etiqueta & tokens_marca
        # Rechaza cualquier etiqueta compuesta casi exclusivamente por tokens de la marca.
        if len(comunes) >= 2 and len(comunes) / len(tokens_etiqueta) >= 0.70:
            return True
    return False

def extract_link(cell):
    if hasattr(cell, "hyperlink") and cell.hyperlink:
        return {"value": "Link", "url": cell.hyperlink.target}
    if isinstance(cell.value, str) and "=HYPERLINK" in cell.value:
        m = re.search(r'=HYPERLINK\("([^"]+)"', cell.value)
        if m: return {"value": "Link", "url": m.group(1)}
    return {"value": cell.value, "url": None}

def extract_link_from_cell(cell):
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target
    return None

def convert_html_entities(text):
    if not isinstance(text, str):
        return text
    text = html.unescape(text)
    html_entities = {
        '&#xF3;': 'ó', '&#xE1;': 'á', '&#xE9;': 'é', '&#xED;': 'í',
        '&#xFA;': 'ú', '&#xF1;': 'ñ', '&#xDC;': 'Ü', '&#xFC;': 'ü',
        '&#xC1;': 'Á', '&#xC9;': 'É', '&#xCD;': 'Í', '&#xD3;': 'Ó',
        '&#xDA;': 'Ú', '&#xD1;': 'Ñ', '&#xC7;': 'Ç', '&#xE7;': 'ç',
    }
    for entity, char in html_entities.items():
        text = text.replace(entity, char)

    def replace_hex_entity(match):
        try:
            return chr(int(match.group(1), 16))
        except Exception:
            return match.group(0)

    def replace_decimal_entity(match):
        try:
            return chr(int(match.group(1)))
        except Exception:
            return match.group(0)

    text = re.sub(r'&#x([0-9A-Fa-f]+);', replace_hex_entity, text)
    text = re.sub(r'&#(\d+);', replace_decimal_entity, text)

    for bad, good in {'\u201c': '"', '\u201d': '"', '\u2018': "'", '\u2019': "'",
                      'Â': '', 'â': '', '€': '', '™': ''}.items():
        text = text.replace(bad, good)
    return text

def clean_text(text):
    if not isinstance(text, str):
        return text
    return convert_html_entities(text).strip()

def clean_cuerpo(text):
    if not isinstance(text, str) or text.strip() == '':
        return text
    text = convert_html_entities(text)
    text = re.sub(r'<br\s*/?>', '\n', text, flags=re.IGNORECASE)
    text = re.sub(r'<[^>]+>', '', text)
    return text.strip()


# ======================================
# FUNCIÓN DE NORMALIZACIÓN DE TÍTULOS (MEJORADA)
# ======================================
def normalize_title_for_comparison(title):
    if not isinstance(title, str): 
        return ""
    
    cleaned = re.sub(r"\s+[\|–—-]\s+[^\|–—-]+$", "", title).strip()
    
    if ":" in cleaned:
        parts = cleaned.split(":", 1)
        suffix = parts[1].strip()
        if len(suffix) >= 10:
            cleaned = suffix
    cleaned = unidecode(cleaned)
    return re.sub(r"\W+", " ", cleaned).lower().strip()


def clean_title_for_output(title):
    return re.sub(r"\s*\|\s*[\w\s]+$", "", str(title)).strip()

def corregir_texto(text):
    if not isinstance(text, str): return text
    text = re.sub(r"(<br>|\[\.\.\.\]|\s+)", " ", text).strip()
    m = re.search(r"[A-ZÁÉÍÓÚÑ]", text)
    if m: text = text[m.start():]
    if text and not text.endswith("..."): text = text.rstrip(".") + "..."
    return text

def normalizar_tipo_medio(tipo_raw):
    if not isinstance(tipo_raw, str): return str(tipo_raw)
    t = unidecode(tipo_raw.strip().lower())
    return {
        'online': 'Internet', 'internet': 'Internet',
        'diario': 'Prensa',
        'am': 'Radio', 'fm': 'Radio', 'radio': 'Radio',
        'aire': 'Televisión', 'cable': 'Televisión', 'tv': 'Televisión',
        'television': 'Televisión', 'televisión': 'Televisión',
        'revista': 'Revistas', 'revistas': 'Revistas',
    }.get(t, str(tipo_raw).strip().title() or "Otro")

def parse_numeric(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if isinstance(val, float) and val.is_integer():
            return int(val)
        return val
    s = str(val).strip()
    if not s:
        return None
    if 'e' in s.lower():
        s = s.replace(',', '.')
    else:
        if ',' in s and '.' in s:
            if s.rfind('.') < s.rfind(','):
                s = s.replace('.', '').replace(',', '.')
            else:
                s = s.replace(',', '')
        elif ',' in s:
            parts = s.split(',')
            if len(parts) > 2 or (len(parts) == 2 and len(parts[1]) == 3 and not s.lower().startswith('0,')):
                s = s.replace(',', '')
            else:
                s = s.replace(',', '.')
        elif '.' in s:
            parts = s.split('.')
            if len(parts) > 2 or (len(parts) == 2 and len(parts[1]) == 3 and not s.lower().startswith('0.')):
                s = s.replace('.', '')
    try:
        f_val = float(s)
        if f_val.is_integer():
            return int(f_val)
        return f_val
    except ValueError:
        return None

def texto_para_embedding(titulo, resumen, max_len=1800):
    t = str(titulo or "").strip()
    r = str(resumen or "").strip()
    return f"{t}. {t}. {t}. {r}"[:max_len]

def _normalizar_mencion(texto: str) -> str:
    """Minúsculas, sin tildes, sin puntuación. utb == UTB, tecnologica == tecnológica."""
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", unidecode(str(texto).lower()))).strip()

def _acronimo_de_nombre(nombre: str) -> str:
    vacias = {"de", "del", "la", "el", "los", "las", "y", "e", "da", "do", "di", "grupo"}
    toks = [t for t in _normalizar_mencion(nombre).split() if t not in vacias]
    if len(toks) < 2:
        return ""
    ac = "".join(t[0] for t in toks)
    return ac if 2 <= len(ac) <= 6 else ""

def _lista_alias(marca, aliases=None):
    nombres = []
    if marca:
        nombres.extend(str(marca).split(";"))
    if isinstance(aliases, str):
        nombres.extend(aliases.split(";"))
    else:
        nombres.extend(str(a) for a in (aliases or []))
    vistos, out = set(), []
    for n in nombres:
        k = _normalizar_mencion(n)
        if k and k not in vistos:
            vistos.add(k)
            out.append(n.strip())
    return out

def _variantes_marca(marca, aliases=None):
    """Formas digitadas + acrónimos (Universidad Tecnológica de Bolívar → utb)."""
    base = _lista_alias(marca, aliases)
    extra = []
    for n in base:
        ac = _acronimo_de_nombre(n)
        if ac:
            extra.append(ac)
    return _lista_alias(";".join(base + extra), None)

def _coincide_nombre_completo(texto: str, nombre: str) -> bool:
    nombre = _normalizar_mencion(nombre)
    if len(nombre) < 2:
        return False
    return bool(re.search(rf"(?<![a-z0-9]){re.escape(nombre)}(?![a-z0-9])", texto))

def _menciona_marca_o_alias(texto: str, marca: str, aliases=None) -> bool:
    normalizado = _normalizar_mencion(texto)
    nombres = _variantes_marca(marca, aliases)
    if any(_coincide_nombre_completo(normalizado, nombre) for nombre in nombres if str(nombre).strip()):
        return True
    tokens_texto = set(normalizado.split())
    vacias = {"de", "del", "la", "el", "los", "las", "y", "grupo"}
    for nombre in nombres:
        tokens_nombre = [t for t in _normalizar_mencion(nombre).split() if len(t) >= 3 and t not in vacias]
        if not tokens_nombre:
            continue
        coincidencias = len(set(tokens_nombre) & tokens_texto)
        if coincidencias >= min(2, len(set(tokens_nombre))) and coincidencias / len(set(tokens_nombre)) >= 0.60:
            return True
        if len(tokens_nombre) == 1:
            token = tokens_nombre[0]
            if len(token) >= 6 and any(
                len(candidate) >= 6 and SequenceMatcher(None, token, candidate).ratio() >= 0.88
                for candidate in tokens_texto
            ):
                return True
        else:
            # Compare each brand/alias token against nearby text tokens. This accepts
            # small spelling differences while still requiring most of the name.
            fuzzy_hits = 0
            for token in set(tokens_nombre):
                if any(
                    candidate == token or (
                        len(token) >= 5 and len(candidate) >= 5
                        and SequenceMatcher(None, token, candidate).ratio() >= 0.86
                    )
                    for candidate in tokens_texto
                ):
                    fuzzy_hits += 1
            required = max(1, int(np.ceil(len(set(tokens_nombre)) * 0.60)))
            if fuzzy_hits >= required and (fuzzy_hits >= 2 or len(set(tokens_nombre)) == 1):
                return True
    return False

def _default_text_column_index(columns, preferred_names, fallback=0):
    """Find common title/summary column spellings without accents or case sensitivity."""
    normalized = [_normalizar_mencion(str(c)).replace("-", " ") for c in columns]
    preferred = [_normalizar_mencion(x).replace("-", " ") for x in preferred_names]
    for wanted in preferred:
        for i, current in enumerate(normalized):
            if current == wanted:
                return i
    for wanted in preferred:
        for i, current in enumerate(normalized):
            if wanted in current or current in wanted:
                return i
    return min(fallback, max(0, len(columns) - 1))

def _safe_filename_part(value):
    cleaned = re.sub(r'[^A-Za-z0-9_-]+', '_', unidecode(str(value or '')).strip())
    return cleaned.strip('_') or 'marca'

def _brand_audit(titulo, resumen, marca, aliases):
    d = extraer_contexto_marca_detallado(titulo, resumen, marca, aliases)
    return d['contexto'], d['coincidencia'], d['origen']

def extraer_contexto_marca(titulo, resumen, marca, aliases=None, ventana=320):
    """Título + resumen si la marca/alias aparece. No recortar tanto como para perder 'ganadores'."""
    titulo = str(titulo or "").strip()
    resumen = str(resumen or "").strip()
    texto = f"{titulo}. {resumen}".strip(" .")
    if not texto or not marca or not _menciona_marca_o_alias(texto, marca, aliases):
        return ""
    partes = re.split(r'(?<=[\.\!\?\n])\s+', texto)
    if len(partes) <= 1:
        partes = re.split(r'\n+', texto)
    hits = [p.strip() for p in partes if p.strip() and _menciona_marca_o_alias(p, marca, aliases)]
    # Siempre incluir el título: ahí suele estar "ganadores" / el hecho.
    bloques = []
    if titulo:
        bloques.append(titulo)
    bloques.extend(hits)
    # Never append the complete summary: sentiment must stay centered on the brand.
    if hits and len(" ".join(hits).split()) < 12:
        all_parts = [p.strip() for p in partes if p.strip()]
        for hit in hits:
            try:
                pos = all_parts.index(hit)
                if pos and all_parts[pos - 1] not in bloques:
                    bloques.insert(0, all_parts[pos - 1])
                elif pos + 1 < len(all_parts) and all_parts[pos + 1] not in bloques:
                    bloques.append(all_parts[pos + 1])
            except ValueError:
                pass
    vistos, out = set(), []
    for h in bloques:
        k = _normalizar_mencion(h)
        if k and k not in vistos:
            vistos.add(k)
            out.append(h)
    return " ".join(out)[:1800] if out else texto[:1800]

def extraer_contexto_marca_detallado(titulo, resumen, marca, aliases=None):
    """Return auditable brand match metadata for sentiment analysis."""
    titulo, resumen = str(titulo or "").strip(), str(resumen or "").strip()
    nombres = _variantes_marca(marca, aliases)
    title_hit = _menciona_marca_o_alias(titulo, marca, aliases)
    summary_hit = _menciona_marca_o_alias(resumen, marca, aliases)
    if not title_hit and not summary_hit:
        return {"contexto": "", "marca_encontrada": "No", "origen": "", "coincidencia": ""}
    origen = ", ".join(x for x, ok in (("Título", title_hit), ("Resumen", summary_hit)) if ok)
    source = f"{titulo}. {resumen}".strip(" .")
    source_norm = _normalizar_mencion(source)
    matched = next((n for n in nombres if _coincide_nombre_completo(source_norm, n)), marca)
    return {
        "contexto": extraer_contexto_marca(titulo, resumen, marca, aliases),
        "marca_encontrada": "Sí", "origen": origen, "coincidencia": matched,
    }

def _validar_etiqueta_completa(etiqueta, titulos_grp=None, resumenes_grp=None, marca="", aliases=None, fallback_fn=None):
    if not etiqueta or etiqueta.strip().lower() in ("sin tema", "varios", "n/a"):
        if fallback_fn: return fallback_fn(titulos_grp or [])
        return "Cobertura informativa general"
    if _frase_esta_completa(etiqueta): return etiqueta
    recortada = _recortar_frase_completa(etiqueta, max_palabras=MAX_PALABRAS_SUBTEMA)
    if _frase_esta_completa(recortada) and len(recortada.split()) >= 2:
        return capitalizar_etiqueta(recortada)
    if titulos_grp and len(titulos_grp) > 0:
        try:
            prompt = (
                f"La frase '{etiqueta}' está incompleta o es genérica. "
                f"Genera una frase temática COMPLETA en español de 3-5 palabras "
                f"con preposición (de/del/para/sobre/en):\n\n"
                + "\n".join(f"  · {t[:120]}" for t in titulos_grp[:4])
                + "\n\nREGLAS: frase nominal con preposición, terminar en sustantivo/adjetivo, "
                f"tildes y ñ correctas. La etiqueta debe explicar el hecho relacionado con '{marca}', "
                "no limitarse al nombre de la institución.\n"
                "CORRECTO: 'Proyecto de terminal de transportes', 'Operación del Canal del Dique'\n"
                "INCORRECTO: 'Terminal transportes', 'Operación canal'\n"
                'JSON: {"subtema":"..."}'
            )
            resp = call_with_retries(
                openai.ChatCompletion.create,
                model=OPENAI_MODEL_CLASIFICACION,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=80,
                temperature=0.1,
                response_format={"type": "json_object"}
            )
            u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
            if u:
                st.session_state['tokens_input'] += (u.get('prompt_tokens') if isinstance(u, dict) else getattr(u, 'prompt_tokens', 0)) or 0
                st.session_state['tokens_output'] += (u.get('completion_tokens') if isinstance(u, dict) else getattr(u, 'completion_tokens', 0)) or 0
            raw = json.loads(resp.choices[0].message.content).get("subtema", "")
            if raw:
                cleaned = limpiar_tema(raw)
                if _frase_esta_completa(cleaned) and len(cleaned.split()) >= 2:
                    return capitalizar_etiqueta(cleaned)
        except:
            pass
    if fallback_fn: return fallback_fn(titulos_grp or [])
    return capitalizar_etiqueta(recortada) if recortada and len(recortada.split()) >= 2 else "Cobertura informativa general"

def dedup_labels(etiquetas, umbral=UMBRAL_DEDUP_LABEL):
    unique = list(dict.fromkeys(etiquetas))
    if len(unique) <= 1:
        return etiquetas
    normed = [string_norm_label(u) for u in unique]
    n = len(unique)
    parent = list(range(n))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra

    def _es_fusion_segura(s1, s2):
        return _etiquetas_compatibles(s1, s2, min_overlap=0.45)

    for i in range(n):
        if not normed[i]: continue
        for j in range(i + 1, n):
            if not normed[j] or find(i) == find(j): continue
            if SequenceMatcher(None, normed[i], normed[j]).ratio() >= max(umbral, 0.88):
                if _es_fusion_segura(normed[i], normed[j]):
                    union(i, j)
                    
    for i in range(n):
        if not normed[i]: continue
        tokens_i = set(normed[i].split())
        if len(tokens_i) < 2: continue
        for j in range(i + 1, n):
            if not normed[j] or find(i) == find(j): continue
            tokens_j = set(normed[j].split())
            if len(tokens_j) < 2: continue
            interseccion = tokens_i & tokens_j
            menor = min(len(tokens_i), len(tokens_j))
            if menor > 0 and len(interseccion) / menor >= 0.78:
                if _es_fusion_segura(normed[i], normed[j]):
                    union(i, j)
                    
    le = get_embeddings_batch(unique)
    vp = [(i, le[i]) for i in range(n) if le[i] is not None]
    if len(vp) >= 2:
        vi, vv = zip(*vp)
        sm = cosine_similarity(np.array(vv))
        for pi in range(len(vi)):
            for pj in range(pi + 1, len(vi)):
                if sm[pi][pj] >= max(umbral, 0.90):
                    if find(vi[pi]) != find(vi[pj]):
                        if _es_fusion_segura(normed[vi[pi]], normed[vi[pj]]):
                            union(vi[pi], vi[pj])

    freq = Counter(etiquetas)
    grupos = defaultdict(list)
    for i in range(n):
        grupos[find(i)].append(i)
    canon = {}
    for root, members in grupos.items():
        cands = [unique[m] for m in members]
        vc = [c for c in cands if c.lower() not in ("sin tema", "varios") and _frase_esta_completa(c)]
        va = [c for c in cands if c.lower() not in ("sin tema", "varios")]
        if vc:
            canon[root] = max(vc, key=lambda c: (freq[c], len(c)))
        elif va:
            best = max(va, key=lambda c: (freq[c], len(c)))
            r = _recortar_frase_completa(best)
            canon[root] = r if _frase_esta_completa(r) else best
        else:
            canon[root] = cands[0]
    lm = {unique[i]: canon[find(i)] for i in range(n)}
    return [capitalizar_etiqueta(lm.get(e, e)) for e in etiquetas]

def _fusionar_subtemas_semanticos(subtemas, textos_por_subtema, marca, aliases, umbral=UMBRAL_FUSION_SUBTEMAS):
    unique_subs = list(dict.fromkeys(subtemas))
    if len(unique_subs) <= 1: return subtemas
    repr_texts = []
    for sub in unique_subs:
        txts = textos_por_subtema.get(sub, [])
        palabras = []
        for t in txts[:20]:
            for w in string_norm_label(str(t)).split():
                if len(w) > 3: palabras.append(w)
        top_kw = " ".join(w for w, _ in Counter(palabras).most_common(10))
        repr_texts.append(f"{sub}. {sub}. {sub}. {top_kw}"[:600])
    emb_repr = get_embeddings_batch(repr_texts)
    valid = [(i, emb_repr[i]) for i in range(len(unique_subs)) if emb_repr[i] is not None]
    if len(valid) < 2: return subtemas
    v_idx, v_emb = zip(*valid)
    sim = cosine_similarity(np.array(v_emb))
    n = len(v_idx)
    parent = list(range(n))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb: parent[rb] = ra

    for i in range(n):
        for j in range(i + 1, n):
            if find(i) == find(j): continue
            sub_i, sub_j = unique_subs[v_idx[i]], unique_subs[v_idx[j]]
            if sim[i][j] >= max(umbral, 0.88) and _grupos_contenido_compatibles(
                textos_por_subtema.get(sub_i, []),
                textos_por_subtema.get(sub_j, []),
                sub_i,
                sub_j,
                min_sim=max(umbral, 0.88),
                min_overlap=0.22,
            ):
                union(i, j)
            
    grupos = defaultdict(list)
    for i in range(n): grupos[find(i)].append(v_idx[i])
    freq = Counter(subtemas)
    lm = {}
    for root, members in grupos.items():
        cands = [unique_subs[m] for m in members]
        if len(cands) == 1:
            lm[cands[0]] = cands[0]
            continue
        vc = [c for c in cands if c.lower() not in ("sin tema", "varios") and _frase_esta_completa(c)]
        best = max(vc, key=lambda c: (freq.get(c, 0), len(c))) if vc else max(cands, key=lambda c: (freq.get(c, 0), len(c)))
        if len(cands) <= 3:
            unified = _unificar_subtemas_llm(cands, textos_por_subtema, marca, aliases)
            if unified and _frase_esta_completa(unified): best = unified
        for c in cands: lm[c] = capitalizar_etiqueta(best)
    return [lm.get(s, s) for s in subtemas]

def _unificar_subtemas_llm(subtemas_a_unificar, textos_por_subtema, marca, aliases):
    subs_str = "\n".join(f"  · {s}" for s in subtemas_a_unificar)
    all_kw = []
    for sub in subtemas_a_unificar:
        for t in textos_por_subtema.get(sub, [])[:5]:
            for w in string_norm_label(str(t)).split():
                if len(w) > 3: all_kw.append(w)
    kw_str = " · ".join(w for w, _ in Counter(all_kw).most_common(8))
    prompt = (
        f"Estos subtemas son variaciones del MISMO tema. "
        f"Genera UN subtema unificado (4-6 palabras) como frase nominal completa:\n\n"
        f"{subs_str}\n\nKeywords: {kw_str}\n\n"
        f"REGLAS: frase coherente vinculada con '{marca}', con preposición (de/del/para/sobre/en), "
        "tildes y ñ correctas, terminar en sustantivo/adjetivo y explicar el hecho, no solo la marca.\n"
        "CORRECTO: 'Regulación de tarifas eléctricas', 'Apertura de nuevas sucursales'\n"
        "INCORRECTO: 'Tarifas energía', 'Apertura sucursales', 'Actividad corporativa'\n"
        'JSON: {"subtema":"..."}'
    )
    try:
        resp = call_with_retries(
            openai.ChatCompletion.create,
            model=OPENAI_MODEL_CLASIFICACION,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=60,
            temperature=0.05,
            response_format={"type": "json_object"}
        )
        u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
        if u:
            st.session_state['tokens_input'] += (u.get('prompt_tokens') if isinstance(u, dict) else getattr(u, 'prompt_tokens', 0)) or 0
            st.session_state['tokens_output'] += (u.get('completion_tokens') if isinstance(u, dict) else getattr(u, 'completion_tokens', 0)) or 0
        raw = json.loads(resp.choices[0].message.content).get("subtema", "")
        if raw: return limpiar_tema(raw)
    except:
        pass
    return None

def get_embeddings_batch(textos, batch_size=100):
    if not textos: return []
    cache = get_embedding_cache()
    resultados, missing = cache.get_many(textos)
    if not missing: return resultados
    mt = [textos[i][:2000] if textos[i] else "" for i in missing]
    for i in range(0, len(mt), batch_size):
        batch = mt[i:i + batch_size]
        bidx = missing[i:i + batch_size]
        try:
            resp = call_with_retries(openai.Embedding.create, input=batch, model=OPENAI_MODEL_EMBEDDING)
            u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
            if u:
                st.session_state['tokens_embedding'] += (u.get('total_tokens') if isinstance(u, dict) else getattr(u, 'total_tokens', 0)) or 0
            for j, d in enumerate(resp["data"]):
                oi = bidx[j]
                emb = d["embedding"]
                resultados[oi] = emb
                cache.put(textos[oi], emb)
        except:
            for j, t in enumerate(batch):
                oi = bidx[j]
                try:
                    r = openai.Embedding.create(input=[t], model=OPENAI_MODEL_EMBEDDING)
                    emb = r["data"][0]["embedding"]
                    resultados[oi] = emb
                    cache.put(textos[oi], emb)
                except:
                    pass
    return resultados

class DSU:
    def __init__(self, n):
        self.p = list(range(n))
        self.rank = [0] * n

    def find(self, i):
        path = []
        while self.p[i] != i:
            path.append(i)
            i = self.p[i]
        for node in path: self.p[node] = i
        return i

    def union(self, i, j):
        ri, rj = self.find(i), self.find(j)
        if ri == rj: return
        if self.rank[ri] < self.rank[rj]: ri, rj = rj, ri
        self.p[rj] = ri
        if self.rank[ri] == self.rank[rj]: self.rank[ri] += 1

    def grupos(self, n):
        c = defaultdict(list)
        for i in range(n): c[self.find(i)].append(i)
        return dict(c)

def agrupar_textos_similares(textos, umbral):
    if not textos: return {}
    embs = get_embeddings_batch(textos)
    valid = [(i, e) for i, e in enumerate(embs) if e is not None]
    if len(valid) < 2: return {}
    idxs, M = zip(*valid)
    labels = AgglomerativeClustering(
        n_clusters=None, distance_threshold=1 - umbral, metric="cosine", linkage="average"
    ).fit(np.array(M)).labels_
    g = defaultdict(list)
    for k, lbl in enumerate(labels): g[lbl].append(idxs[k])
    return dict(enumerate(g.values()))

def agrupar_por_titulo_similar(titulos):
    gid, grupos, used = 0, {}, set()
    norm = [normalize_title_for_comparison(t) for t in titulos]
    for i in range(len(norm)):
        if i in used or not norm[i]: continue
        grp = [i]
        used.add(i)
        for j in range(i + 1, len(norm)):
            if j in used or not norm[j]: continue
            if SequenceMatcher(None, norm[i], norm[j]).ratio() >= SIMILARITY_THRESHOLD_TITULOS:
                grp.append(j)
                used.add(j)
        if len(grp) >= 2:
            grupos[gid] = list(set(grp))
            gid += 1
    return grupos

def seleccionar_representante(indices, textos):
    embs = get_embeddings_batch([textos[i] for i in indices])
    validos = [(indices[k], e) for k, e in enumerate(embs) if e is not None]
    if not validos: return indices[0], textos[indices[0]]
    idxs, M = zip(*validos)
    centro = np.mean(M, axis=0, keepdims=True)
    best = int(np.argmax(cosine_similarity(np.array(M), centro)))
    return idxs[best], textos[idxs[best]]

def construir_grupos_consistentes(titulos, resumenes):
    """Agrupa republicaciones y notas equivalentes con criterios conservadores."""
    titulos = [str(x or "") for x in titulos]
    resumenes = [str(x or "") for x in resumenes]
    textos = [texto_para_embedding(t, r) for t, r in zip(titulos, resumenes)]
    n = len(textos)
    dsu = DSU(n)
    embs = get_embeddings_batch(textos)
    norm = [normalize_title_for_comparison(t) for t in titulos]

    # Bloqueo por palabras distintivas para evitar una comparación O(n²) completa.
    indice = defaultdict(set)
    for i, titulo in enumerate(norm):
        for token in _tokens_distintivos(titulo):
            indice[token].add(i)
    pares = set()
    for idxs in indice.values():
        if len(idxs) > 100:
            continue
        orden = sorted(idxs)
        pares.update((orden[a], orden[b]) for a in range(len(orden)) for b in range(a + 1, len(orden)))

    for i, j in pares:
        if _hay_conflicto_accion(textos[i], textos[j]):
            continue
        title_sim = SequenceMatcher(None, norm[i], norm[j]).ratio() if norm[i] and norm[j] else 0.0
        overlap = _overlap_distintivo(textos[i], textos[j])
        semantic = 0.0
        if embs[i] is not None and embs[j] is not None:
            semantic = cosine_similarity(
                np.array(embs[i]).reshape(1, -1), np.array(embs[j]).reshape(1, -1)
            )[0][0]
        if title_sim >= SIMILARITY_THRESHOLD_TITULOS or (semantic >= SIMILARITY_THRESHOLD_TONO and overlap >= 0.45):
            dsu.union(i, j)
    return dsu.grupos(n)

def aplicar_consistencia_grupos(df, titulo_col, resumen_col,
                                tono_col="Tono IA", tema_col="Tema", subtema_col="Subtema"):
    """Asigna Grupo noticia como overlay. No sobrescribe Tono IA / Tema / Subtema."""
    if df.empty:
        return df
    grupos = construir_grupos_consistentes(df[titulo_col].fillna(''), df[resumen_col].fillna(''))
    df = df.copy()
    df["Grupo noticia"] = ""
    for numero, idxs in enumerate(grupos.values(), start=1):
        gid = f"G{numero:05d}"
        for i in idxs:
            df.at[df.index[i], "Grupo noticia"] = gid
        # Keep existing labels even when members disagree. Majority vote was
        # stamping a generic Subtema onto notes that must stay specific.
    if subtema_col in df.columns:
        df[subtema_col] = df[subtema_col].apply(
            lambda x: capitalizar_etiqueta(_recortar_frase_completa(str(x), MAX_PALABRAS_SUBTEMA))
            if str(x).strip().lower() not in {"", "nan", "n/a", "-"} else x
        )
    return df


# ======================================
# TONO (Sistema Reputacional por IA)
# ======================================
class ClasificadorTono:
    def __init__(self, marca, aliases):
        nombres = _variantes_marca(marca, aliases)
        self.marca = nombres[0] if nombres else str(marca or "").strip()
        self.aliases = [n for n in nombres[1:] if n]
        self._all_names = [self.marca] + self.aliases

    def _menciona_marca(self, texto):
        return _menciona_marca_o_alias(texto, self.marca, self.aliases)

    async def _clasificar_llm(self, texto, sem, contexto_marca=""):
        async with sem:
            eval_txt = (contexto_marca or texto or "").strip()
            if not eval_txt or not self._menciona_marca(eval_txt):
                return {"tono": "Neutro"}

            aliases_str = f" (también conocida como: {', '.join(self.aliases)})" if self.aliases else ""
            prompt = (
                f"Eres un experto analista en Relaciones Públicas y Gestión de Reputación. "
                f"Evalúa el impacto reputacional DIRECTO sobre la marca '{self.marca}'{aliases_str}.\n\n"
                f"El tono GENERAL de la noticia NO importa. Si el artículo es neutro o habla de otro actor, "
                f"pero la mención a '{self.marca}' es favorable, el tono es Positivo. "
                f"Si '{self.marca}' gana, es premiada, finalista, reconocida o se alza como ganadora "
                f"(aunque aparezca junto a otras instituciones), el tono es Positivo. "
                f"Estar en una LISTA DE GANADORES no es Neutro. "
                f"Si el artículo es positivo o trágico a nivel país/sector, pero '{self.marca}' queda "
                f"criticada, demandada o cuestionada, el tono es Negativo.\n\n"
                f"TEXTO CENTRADO EN LA MARCA:\n{eval_txt[:1600]}\n\n"
                f"REGLAS DE CLASIFICACIÓN ESTRICTAS:\n"
                f"🔴 NEGATIVO: un hecho perjudica, cuestiona o expone directamente a '{self.marca}' "
                f"(demandas, multas, fraudes, fallas propias, quejas, investigaciones, pérdidas o retiro de productos).\n"
                f"🟢 POSITIVO: el hecho acredita directamente un logro, mejora o aporte verificable de '{self.marca}' "
                f"(premio, crecimiento, lanzamiento exitoso, inversión realizada, innovación, expansión o reconocimiento).\n"
                f"⚪ NEUTRO: La marca se menciona SIN impacto a su imagen. Ejemplos:\n"
                f"  - La noticia habla de una crisis del sector/país, pero la marca solo es mencionada informando o adaptándose.\n"
                f"  - Se menciona a la marca de paso, sin rol (no aplica si es ganadora, premiada o protagonista).\n"
                f"  - Una persona, autoridad, proveedor o tercero es quien recibe el efecto positivo o negativo.\n"
                f"  - Emite un comunicado regular sin evidencia de crisis ni logro relevante.\n"
                f"  - Critica, denuncia o advierte sobre un problema de terceros o del sector; la crítica de la marca NO es una crítica contra la marca.\n\n"
                f"⚠️ ATENCIÓN: Ignora el tono del sector o de terceros. Evalúa ÚNICAMENTE cómo el hecho afecta "
                f"la reputación corporativa de '{self.marca}': mejora (Positivo), empeora (Negativo) o no cambia (Neutro).\n\n"
                f'Responde ÚNICAMENTE con JSON: {{"tono":"Positivo|Negativo|Neutro", '
                f'"confianza":"Alta|Media|Baja", "justificacion":"explicación concreta de máximo 35 palabras"}}'
            )

            try:
                resp = await acall_with_retries(
                    openai.ChatCompletion.acreate,
                    model=OPENAI_MODEL_CLASIFICACION,
                    messages=[{"role": "user", "content": prompt}],
                    max_tokens=100,
                    temperature=0.0,
                    response_format={"type": "json_object"}
                )
                
                u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
                if u:
                    st.session_state['tokens_input'] += (u.get('prompt_tokens') if isinstance(u, dict) else getattr(u, 'prompt_tokens', 0)) or 0
                    st.session_state['tokens_output'] += (u.get('completion_tokens') if isinstance(u, dict) else getattr(u, 'completion_tokens', 0)) or 0
                
                resultado = json.loads(resp.choices[0].message.content)
                tono = str(resultado.get("tono", "Neutro")).strip().title()
                
                tono = tono if tono in ("Positivo", "Negativo", "Neutro") else "Neutro"
                confianza = str(resultado.get("confianza", "Media")).strip().title()
                if confianza not in ("Alta", "Media", "Baja"):
                    confianza = "Media"
                return {"tono": tono, "confianza": confianza,
                        "justificacion": str(resultado.get("justificacion", "")).strip()[:400],
                        "evidencia": eval_txt[:1800]}
            except Exception as e:
                return {"tono": "Neutro", "confianza": "Baja", "justificacion": "Error de clasificación", "evidencia": eval_txt[:1800]}

    async def procesar_lote_async(self, textos, pbar, resumenes, titulos):
        n = len(textos)
        txts = textos.tolist()
        pbar.progress(0.05, "Agrupando noticias para análisis de tono...")
        
        txts_emb = [texto_para_embedding(str(titulos.iloc[i]), str(resumenes.iloc[i])) for i in range(n)]
        dsu = DSU(n)
        
        embs = get_embeddings_batch(txts_emb)
        candidatos = agrupar_textos_similares(txts_emb, SIMILARITY_THRESHOLD_TONO)
        candidatos.update({len(candidatos) + k: v for k, v in agrupar_por_titulo_similar(titulos.tolist()).items()})
        for idxs in candidatos.values():
            for pos, i in enumerate(idxs):
                for j in idxs[pos + 1:]:
                    ti, tj = normalize_title_for_comparison(titulos.iloc[i]), normalize_title_for_comparison(titulos.iloc[j])
                    titulo_casi_igual = SequenceMatcher(None, ti, tj).ratio() >= 0.96
                    contenido_casi_igual = (
                        embs[i] is not None and embs[j] is not None
                        and cosine_similarity(np.array(embs[i]).reshape(1, -1), np.array(embs[j]).reshape(1, -1))[0][0] >= SIMILARITY_THRESHOLD_TONO
                        and _overlap_distintivo(txts_emb[i], txts_emb[j]) >= 0.45
                    )
                    if (titulo_casi_igual or contenido_casi_igual) and not _hay_conflicto_accion(txts_emb[i], txts_emb[j]):
                        dsu.union(i, j)
                
        grupos = dsu.grupos(n)
        contextos = [
            extraer_contexto_marca(str(titulos.iloc[i]), str(resumenes.iloc[i]), self.marca, self.aliases)
            for i in range(n)
        ]
        reps = {}
        for cid, idxs in grupos.items():
            con_marca = [i for i in idxs if contextos[i]]
            if con_marca:
                ri, _ = seleccionar_representante(con_marca, contextos)
                reps[cid] = (ri, contextos[ri])
            else:
                reps[cid] = (idxs[0], "")
        
        sem = asyncio.Semaphore(CONCURRENT_REQUESTS)
        cids = list(reps.keys())
        
        async def _clasificar_con_cid(cid):
            _idx, ctx = reps[cid]
            if not ctx:
                return cid, {"tono": "Neutro"}
            return cid, await self._clasificar_llm(txts[_idx], sem, contexto_marca=ctx)

        tasks = [_clasificar_con_cid(c) for c in cids]
        rpg = {}
        
        for i, f in enumerate(asyncio.as_completed(tasks)):
            cid, r = await f
            rpg[cid] = r
            pbar.progress(0.1 + 0.85 * (i + 1) / len(tasks), f"Evaluando Reputación {i + 1}/{len(tasks)}")
            
        final = [None] * n
        
        for cid, idxs in grupos.items():
            r = rpg.get(cid, {"tono": "Neutro"})
            for i in idxs: final[i] = r

        tonos = [f["tono"] if f else "Neutro" for f in final]
        tonos = _propagar_tono_equivalentes(tonos, titulos.tolist(), resumenes.tolist())
        final = [{"tono": t} for t in tonos]
            
        pbar.progress(1.0, "Análisis de Tono completado")
        return final

def _propagar_tono_equivalentes(tonos, titulos, resumenes):
    """Noticias equivalentes (cualquier marca): si una es Positivo/Negativo y otra Neutro, se alinean."""
    n = len(tonos)
    if n < 2:
        return list(tonos)
    out = list(tonos)
    norm_t = [_normalizar_mencion(normalize_title_for_comparison(t) or t) for t in titulos]
    norm_r = [_normalizar_mencion(str(r)[:320]) for r in resumenes]
    combos = [_normalizar_mencion(f"{titulos[i]} {str(resumenes[i])[:320]}") for i in range(n)]
    dsu = DSU(n)
    for i in range(n):
        for j in range(i + 1, n):
            sim_t = SequenceMatcher(None, norm_t[i], norm_t[j]).ratio() if norm_t[i] and norm_t[j] else 0.0
            sim_r = SequenceMatcher(None, norm_r[i], norm_r[j]).ratio() if norm_r[i] and norm_r[j] else 0.0
            ov = _overlap_distintivo(combos[i], combos[j])
            mismo = (
                sim_t >= 0.80
                or sim_r >= 0.78
                or (sim_t >= 0.62 and sim_r >= 0.58)
                or (ov >= 0.50 and (sim_t >= 0.55 or sim_r >= 0.55))
            )
            if mismo and not _hay_conflicto_accion(combos[i], combos[j]):
                dsu.union(i, j)
    for idxs in dsu.grupos(n).values():
        if len(idxs) < 2:
            continue
        vals = [out[i] for i in idxs]
        if "Positivo" in vals and "Negativo" in vals:
            continue
        if "Positivo" in vals:
            canon = "Positivo"
        elif "Negativo" in vals:
            canon = "Negativo"
        else:
            continue
        for i in idxs:
            if out[i] in ("Neutro", "N/A", "", "Nan"):
                out[i] = canon
    return out

def analizar_tono_con_pkl(textos, pkl_file, titulos=None, resumenes=None, marca="", aliases=None):
    try:
        pipeline = joblib.load(pkl_file)
        TM = {
            1: "Positivo", "1": "Positivo", "positivo": "Positivo", "Positivo": "Positivo",
            0: "Neutro", "0": "Neutro", "neutro": "Neutro", "Neutro": "Neutro",
            -1: "Negativo", "-1": "Negativo", "negativo": "Negativo", "Negativo": "Negativo",
        }
        def _norm_pred(p):
            if p in TM: return TM[p]
            s = str(p).strip()
            return TM.get(s, TM.get(s.title(), s.title() if s.title() in ("Positivo", "Negativo", "Neutro") else "Neutro"))

        if marca and titulos is not None and resumenes is not None:
            titulos = list(titulos)
            resumenes = list(resumenes)
            n = len(titulos)
            snippets, flags = [], []
            for i in range(n):
                ctx = extraer_contexto_marca(titulos[i], resumenes[i], marca, aliases)
                if ctx:
                    tit = str(titulos[i] or "").strip()
                    snippet = ctx if (tit and ctx.lower().startswith(tit[:20].lower())) else (f"{tit}. {ctx}" if tit else ctx)
                    snippets.append(snippet[:1800])
                    flags.append(True)
                else:
                    snippets.append("")
                    flags.append(False)
            result = [{"tono": "Neutro"}] * n
            idx_pred = [i for i, f in enumerate(flags) if f]
            if idx_pred:
                preds = pipeline.predict([snippets[i] for i in idx_pred])
                for i, p in zip(idx_pred, preds):
                    result[i] = {"tono": _norm_pred(p)}
            if titulos is not None and resumenes is not None:
                tonos = _propagar_tono_equivalentes([r["tono"] for r in result], list(titulos), list(resumenes))
                return [{"tono": t} for t in tonos]
            return result
        preds = [{"tono": _norm_pred(p)} for p in pipeline.predict(textos)]
        if titulos is not None and resumenes is not None:
            tonos = _propagar_tono_equivalentes([r["tono"] for r in preds], list(titulos), list(resumenes))
            return [{"tono": t} for t in tonos]
        return preds
    except Exception as e:
        st.error(f"Error pkl tono: {e}")
        return None

def analizar_temas_con_pkl(textos, pkl_file):
    try:
        pipeline = joblib.load(pkl_file)
        predicciones = pipeline.predict(textos)
        return [str(p).strip() for p in predicciones]
    except Exception as e:
        st.error(f"Error pkl temas: {e}")
        return None

# ======================================
# SUBTEMAS
# ======================================
class ClasificadorSubtema:
    def __init__(self, marca, aliases):
        nombres = _variantes_marca(marca, aliases)
        self.marca = nombres[0] if nombres else str(marca or "")
        self.aliases = nombres[1:]
        self._cache = {}
        self._umbrales: dict = {}

    def _paso1(self, titulos, resumenes, dsu):
        def nt(t, n):
            return ' '.join(re.sub(r'[^a-z0-9\s]', '', unidecode(str(t).lower())).split()[:n])
        bt, br = defaultdict(list), defaultdict(list)
        for i, (ti, re_) in enumerate(zip(titulos, resumenes)):
            a, b = nt(ti, 40), nt(re_, 15)
            if a: bt[hashlib.md5(a.encode()).hexdigest()].append(i)
            b = nt(re_, 120)
            if len(b.split()) >= 25: br[hashlib.md5(b.encode()).hexdigest()].append(i)
        for bk in (bt, br):
            for idxs in bk.values():
                for j in idxs[1:]: dsu.union(idxs[0], j)

    def _paso2(self, titulos, dsu):
        norm = [normalize_title_for_comparison(t) for t in titulos]
        n = len(norm)
        for i in range(n):
            if not norm[i]: continue
            for j in range(i + 1, n):
                if not norm[j] or dsu.find(i) == dsu.find(j): continue
                ratio = SequenceMatcher(None, norm[i], norm[j]).ratio()
                comparte_asunto = _overlap_distintivo(norm[i], norm[j]) >= 0.40
                if ratio >= SIMILARITY_THRESHOLD_TITULOS and comparte_asunto and not _hay_conflicto_accion(norm[i], norm[j]):
                    dsu.union(i, j)

    def _paso2b_keywords(self, titulos, dsu, ae):
        sim_min = self._umbrales.get('sim_minima_keywords', SIM_MINIMA_KEYWORDS_RARAS)
        stop = {
            'el','la','los','las','un','una','unos','unas','de','del','al',
            'en','con','por','para','que','se','su','sus','es','son','fue',
            'como','mas','pero','sin','sobre','entre','tras','esta','este',
            'esto','hay','ser','han','ha','ya','muy','otro','otra','otros',
            'otras','todo','toda','todos','todas','puede','desde','hasta',
            'donde','cuando','quien','cual','cada','nos','les','ante','bajo',
            'nueva','nuevo','nuevos','nuevas','forma','hace','asi','sera',
            'segun','tiene','fueron','sido','hacer','dice','dijo','tambien',
        }
        titulo_words = []
        for t in titulos:
            ws = set()
            for w in re.findall(r'[a-z]+', unidecode(str(t).lower())):
                if len(w) >= 5 and w not in stop: ws.add(w)
            titulo_words.append(ws)
        word_freq = Counter()
        for ws in titulo_words:
            for w in ws: word_freq[w] += 1
        n = len(titulos)
        max_freq = max(2, int(n * 0.03))
        rare_index = defaultdict(list)
        for i, ws in enumerate(titulo_words):
            for w in ws:
                if 2 <= word_freq[w] <= max_freq: rare_index[w].append(i)
        for idxs in rare_index.values():
            for a in range(len(idxs)):
                for b in range(a + 1, len(idxs)):
                    ia, ib = idxs[a], idxs[b]
                    if dsu.find(ia) == dsu.find(ib): continue
                    ea, eb = ae[ia], ae[ib]
                    if ea is None or eb is None: continue
                    sim = cosine_similarity(
                        np.array(ea).reshape(1, -1),
                        np.array(eb).reshape(1, -1)
                    )[0][0]
                    if sim >= sim_min and not _hay_conflicto_accion(str(titulos[ia]), str(titulos[ib])):
                        dsu.union(ia, ib)

    def _paso3(self, et, ae, dsu, pbar, ps):
        umbral_cluster = max(self._umbrales.get('subtema', UMBRAL_SUBTEMA), 0.82)
        sim_min = max(self._umbrales.get('sim_minima_agrupacion', SIM_MINIMA_AGRUPACION_SUBTEMA), 0.90)
        n = len(et)
        if n < 2: return

        def _puede_unir(i, j):
            if _hay_conflicto_accion(et[i], et[j]):
                return False
            if _overlap_distintivo(et[i], et[j]) >= 0.30:
                return True
            return SequenceMatcher(
                None,
                normalize_title_for_comparison(et[i]),
                normalize_title_for_comparison(et[j])
            ).ratio() >= 0.96

        B = 500
        if n <= B:
            pbar.progress(ps, "Clustering semántico...")
            ok = [(k, e) for k, e in enumerate(ae) if e is not None]
            if len(ok) < 2: return
            io_, M = zip(*ok)
            sim_matrix = cosine_similarity(np.array(M))
            linkage = 'complete' if n <= 10 else 'average'
            labels = AgglomerativeClustering(
                n_clusters=None, distance_threshold=1 - umbral_cluster,
                metric='precomputed', linkage=linkage
            ).fit(1 - sim_matrix).labels_
            g = defaultdict(list)
            for k, lbl in enumerate(labels): g[lbl].append(io_[k])
            for cl in g.values():
                if len(cl) < 2: continue
                vecs = np.array([ae[i] for i in cl if ae[i] is not None])
                if len(vecs) < 2: continue
                centroid = np.mean(vecs, axis=0)
                sims_al_centroid = cosine_similarity(vecs, centroid.reshape(1, -1)).flatten()
                todos_ok = all(s >= sim_min for s in sims_al_centroid)
                if todos_ok:
                    for j in cl[1:]:
                        if _puede_unir(cl[0], j):
                            dsu.union(cl[0], j)
                else:
                    mejor_idx = int(np.argmax(sims_al_centroid))
                    repr_vec = np.array(ae[cl[mejor_idx]]).reshape(1, -1)
                    for k_local, i_global in enumerate(cl):
                        if ae[i_global] is None: continue
                        sim_vs_repr = cosine_similarity(
                            np.array(ae[i_global]).reshape(1, -1), repr_vec
                        )[0][0]
                        if sim_vs_repr >= sim_min and _puede_unir(cl[mejor_idx], i_global):
                            dsu.union(cl[mejor_idx], i_global)
            pbar.progress(ps + 0.18, "Clustering completado")
            return

        tb = max(1, (n + B - 1) // B)
        for bn_, bs in enumerate(range(0, n, B)):
            bi = list(range(bs, min(bs + B, n)))
            ok = [(idx, ae[idx]) for idx in bi if ae[idx] is not None]
            if len(ok) < 2: continue
            io_, M = zip(*ok)
            sim_matrix = cosine_similarity(np.array(M))
            labels = AgglomerativeClustering(
                n_clusters=None, distance_threshold=1 - umbral_cluster,
                metric='precomputed', linkage='average'
            ).fit(1 - sim_matrix).labels_
            g = defaultdict(list)
            for k, lbl in enumerate(labels): g[lbl].append(io_[k])
            for cl in g.values():
                if len(cl) < 2: continue
                vecs = np.array([ae[i] for i in cl if ae[i] is not None])
                if len(vecs) < 2: continue
                centroid = np.mean(vecs, axis=0)
                sims = cosine_similarity(vecs, centroid.reshape(1, -1)).flatten()
                mejor_idx = int(np.argmax(sims))
                repr_vec = np.array(ae[cl[mejor_idx]]).reshape(1, -1)
                for k_local, i_global in enumerate(cl):
                    if ae[i_global] is None: continue
                    s = cosine_similarity(np.array(ae[i_global]).reshape(1, -1), repr_vec)[0][0]
                    if s >= sim_min and _puede_unir(cl[mejor_idx], i_global):
                        dsu.union(cl[mejor_idx], i_global)
            pbar.progress(ps + 0.15 * (bn_ + 1) / tb, f"Clustering {bn_ + 1}/{tb}...")

        pbar.progress(ps + 0.16, "Unificando...")
        usar_fusion = self._umbrales.get('usar_fusion_iterativa', True)
        if usar_fusion: self._fusion(et, ae, dsu, pbar, ps + 0.16)

    def _fusion(self, textos, ae, dsu, pbar, ps):
        n = len(textos)
        umbral_inter = self._umbrales.get('fusion_intergrupo', UMBRAL_FUSION_INTERGRUPO)
        max_iter = self._umbrales.get('max_iter_fusion', MAX_ITER_FUSION)
        sim_min = self._umbrales.get('sim_minima_agrupacion', SIM_MINIMA_AGRUPACION_SUBTEMA)
        for it in range(max_iter):
            grupos = dsu.grupos(n)
            if len(grupos) < 2: break
            centroids, vg = [], []
            for gid, idxs in grupos.items():
                vecs = [ae[i] for i in idxs[:50] if ae[i] is not None]
                if vecs:
                    centroids.append(np.mean(vecs, axis=0))
                    vg.append(gid)
            if len(vg) < 2: break
            sim = cosine_similarity(np.array(centroids))
            umbral_efectivo = max(umbral_inter, sim_min)
            pairs = sorted(
                [(sim[i][j], i, j) for i in range(len(vg)) for j in range(i + 1, len(vg))
                 if sim[i][j] >= umbral_efectivo], reverse=True
            )
            fus = 0
            for _, i, j in pairs:
                ri, rj = grupos[vg[i]][0], grupos[vg[j]][0]
                if dsu.find(ri) != dsu.find(rj):
                    textos_i = [textos[k] for k in grupos[vg[i]][:20]]
                    textos_j = [textos[k] for k in grupos[vg[j]][:20]]
                    if _grupos_contenido_compatibles(
                        textos_i,
                        textos_j,
                        "",
                        "",
                        min_sim=umbral_efectivo,
                        min_overlap=0.16,
                    ):
                        dsu.union(ri, rj)
                        fus += 1
            pbar.progress(min(ps + 0.04 * (it + 1), 0.52), f"Fusión {it + 1}: {fus}")
            if fus == 0: break

    def _extraer_keywords_titulos(self, titulos_grp: list, top_n: int = 6) -> list:
        palabras = []
        for t in titulos_grp[:10]:
            for w in string_norm_label(t).split():
                if len(w) > 3: palabras.append(w)
        return [w for w, _ in Counter(palabras).most_common(top_n)]

    def _generar_etiqueta(self, textos_grp, titulos_grp, resumenes_grp, subtemas_existentes=None, evitar_etiqueta=None):
        tn = sorted(set(normalize_title_for_comparison(t) for t in titulos_grp if t))
        existentes_key = "|".join(sorted(string_norm_label(s) for s in (subtemas_existentes or []))[:20])
        evitar_key = string_norm_label(evitar_etiqueta) if evitar_etiqueta else ""
        ck = hashlib.md5(("|".join(tn[:12]) + f"#{len(titulos_grp)}#{existentes_key}#{evitar_key}").encode()).hexdigest()
        if ck in self._cache: return self._cache[ck]

        tm = list(dict.fromkeys(str(t)[:130] for t in titulos_grp if pd.notna(t) and str(t).strip() and str(t).strip().lower() != 'nan'))[:6]
        rm = [str(r)[:200] for r in resumenes_grp[:3] if r and len(str(r)) > 20]

        kw_list = self._extraer_keywords_titulos(titulos_grp, top_n=8)
        palabras_res = []
        for r in resumenes_grp[:5]:
            for w in string_norm_label(str(r)).split():
                if len(w) > 4: palabras_res.append(w)
        kw_res = [w for w, _ in Counter(palabras_res).most_common(4)
                  if w not in {unidecode(k.lower()) for k in kw_list}]
        kw_todos = kw_list + kw_res
        kw = ", ".join(kw_todos[:10])

        ctx_resumenes = (
            "\nRESÚMENES (para contexto):\n"
            + "\n".join(f"  · {r}" for r in rm)
        ) if rm else ""

        if len(kw_list) >= 3:
            ejemplo_dinamico = (
                f"'{kw_list[0].title()} de {kw_list[1].title()}' o "
                f"'{kw_list[0].title()} del {kw_list[2].title()}'"
            )
        elif len(kw_list) >= 2:
            ejemplo_dinamico = f"'{kw_list[0].title()} de {kw_list[1].title()}'"
        elif len(kw_list) == 1:
            ejemplo_dinamico = f"'{kw_list[0].title()} en la región'"
        else:
            ejemplo_dinamico = "'Proyecto de terminal de transportes'"

        lista_existentes = ""
        if subtemas_existentes and len(subtemas_existentes) > 0:
            lista_existentes = (
                "\n\nSUBTEMAS YA CREADOS (ÚSALOS SI APLICAN EXACTAMENTE):\n" +
                ", ".join(f"'{s}'" for s in subtemas_existentes[:15]) +
                "\nREGLA: Si este grupo de noticias trata EXACTAMENTE del mismo tema que uno de los subtemas ya creados, responde con ese subtema. Si es un tema distinto, crea uno nuevo."
            )
        if evitar_etiqueta:
            lista_existentes += (
                f"\nNO uses '{evitar_etiqueta}': este grupo es un evento distinto, genera un subtema nuevo y específico."
            )

        prompt = (
            f"Eres analista de reputación de la marca principal '{self.marca}'. "
            "Genera UN subtema periodístico (3-5 palabras) que sea una FRASE NOMINAL "
            "— sin sujeto ni verbo conjugado — para este grupo de noticias.\n\n"
            "TÍTULOS:\n" + "\n".join(f"  · {t}" for t in tm)
            + ctx_resumenes
            + f"\n\nPALABRAS CLAVE: {kw}"
            + lista_existentes
            + "\n\nREGLAS OBLIGATORIAS:\n"
            + f"  0. Usa solo hechos vinculados con '{self.marca}' o sus alias; valida la mención en título o resumen.\n"
            + "     El subtema debe describir qué ocurre con la marca (no solo repetir su nombre).\n"
            "  1. FRASE NOMINAL PURA: empieza con sustantivo, usa preposición para unir conceptos.\n"
            "     NUNCA empieces con cargo/persona ('Alcalde', 'Gobernador', 'Ministro').\n"
            "     NUNCA incluyas verbo conjugado ('presenta', 'anuncia', 'lanza', 'inaugura').\n"
            f"     CORRECTO: {ejemplo_dinamico}\n"
            "     INCORRECTO: 'Alcalde presenta proyecto terminal', "
            "'Gobernador anuncia inversión', 'Alcaldía lanza plan'\n"
            "  2. USA preposiciones (de, del, para, sobre, en, por) para conectar concepts.\n"
            "  3. SÉ ESPECÍFICO: describe el asunto real, no el actor.\n"
            "  4. Ciudades y regiones SÍ pueden aparecer si son relevantes al tema.\n"
            "  5. Puedes usar la marca completa si aporta claridad y cabe en cinco palabras. Tildes y ñ correctas.\n\n"
            "EJEMPLOS CORRECTOS: 'Proyecto de terminal de transportes', "
            "'Operación del Canal del Dique', 'Plan de infraestructura vial', "
            "'Regulación de tarifas eléctricas', 'Inversión en salud pública'\n"
            "EJEMPLOS INCORRECTOS: 'Alcalde presenta proyecto', 'Gobernador lanza plan', "
            "'Tarifas energía', 'Gestión corporativa', 'Actividad legislativa'\n\n"
            'JSON: {"subtema":"..."}'
        )

        _VERBOS_FRASES = re.compile(
            r'\b(presenta|presentan|anuncia|anuncian|lanza|lanzan|inaugura|inauguran|'
            r'realiza|realizan|desarrolla|desarrollan|ejecuta|ejecutan|gestiona|gestionan|'
            r'impulsa|impulsan|promueve|promueven|lidera|lideran|encabeza|encabezan|'
            r'aprueba|aprueban|firma|firman|suscribe|suscriben|invierte|invierten|'
            r'construye|construyen|instala|instalan|entrega|entregan|recibe|reciben|'
            r'solicita|solicitan|visita|visitan|atiende|atienden|destaca|destacan|'
            r'señala|señalan|indica|indican|expresa|expresan|afirma|afirman|'
            r'propone|proponen|pide|piden|exige|exigen|apoya|apoyan|'
            r'informa|informan|reporta|reportan|advierte|advierten)\b',
            re.IGNORECASE
        )

        def _tiene_verbo_conjugado(s): return bool(_VERBOS_FRASES.search(s))

        try:
            resp = call_with_retries(
                openai.ChatCompletion.create,
                model=OPENAI_MODEL_CLASIFICACION,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=60,
                temperature=0.0,
                response_format={"type": "json_object"}
            )
            u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
            if u:
                st.session_state['tokens_input'] += (u.get('prompt_tokens') if isinstance(u, dict) else getattr(u, 'prompt_tokens', 0)) or 0
                st.session_state['tokens_output'] += (u.get('completion_tokens') if isinstance(u, dict) else getattr(u, 'completion_tokens', 0)) or 0

            raw = json.loads(resp.choices[0].message.content).get("subtema", "Varios")
            et = limpiar_tema(raw)

            if not et or et.strip().lower() == "sin tema":
                et = self._refinar(tm, kw, rm, forzar_preposicion=True)
            if _tiene_verbo_conjugado(et):
                et = self._refinar(tm, kw, rm, forzar_preposicion=True, prohibir_verbos=True)

            def _es_robotico(s):
                palabras = s.split()
                if len(palabras) <= 3:
                    nexos = {"de", "del", "para", "sobre", "en", "con", "por",
                             "ante", "hacia", "entre", "sin", "al", "las", "los",
                             "una", "uno", "que", "como", "y", "o", "a", "e", "u"}
                    tiene_nexo = any(unidecode(p.lower()) in nexos for p in palabras[1:])
                    if not tiene_nexo: return True
                return False

            genericas = {"gestión", "gestion", "actividades", "acciones", "noticias",
                         "información", "informacion", "eventos", "varios", "sin tema",
                         "actividad corporativa", "gestion corporativa"}
            es_gen = string_norm_label(et) in {string_norm_label(g) for g in genericas}
            es_solo_marca = _es_nombre_o_fragmento_marca(et, self.marca, self.aliases)
            es_rob = _es_robotico(et)

            if es_gen or es_solo_marca or es_rob or len(et.split()) < 3:
                et = self._refinar(tm, kw, rm, forzar_preposicion=True)

            if not _validar_estructura_subtema(et):
                et = self._refinar(tm, kw, rm, forzar_preposicion=True)
                if not _validar_estructura_subtema(et):
                    et = self._fallback(titulos_grp)

            et = _validar_etiqueta_completa(
                et, titulos_grp=titulos_grp, resumenes_grp=resumenes_grp,
                marca=self.marca, aliases=self.aliases, fallback_fn=self._fallback
            )
            if _es_nombre_o_fragmento_marca(et, self.marca, self.aliases):
                et = self._refinar(tm, kw, rm, forzar_preposicion=True, prohibir_verbos=True)
            if _es_nombre_o_fragmento_marca(et, self.marca, self.aliases):
                et = self._fallback(titulos_grp)
        except:
            et = self._fallback(titulos_grp)

        et = capitalizar_etiqueta(et)
        self._cache[ck] = et
        return et

    def _refinar(self, titulos, kw, resumenes=None, forzar_preposicion=False, prohibir_verbos=False):
        ctx = ("\nContexto de resúmenes: " + " | ".join(r[:100] for r in resumenes[:3])) if resumenes else ""
        kw_parts = [w.strip() for w in kw.split(",") if w.strip()]

        if len(kw_parts) >= 3:
            ej_bueno = f"'{kw_parts[0].title()} de {kw_parts[1].title()}', '{kw_parts[0].title()} en {kw_parts[2].title()}'"
        elif len(kw_parts) >= 2:
            ej_bueno = f"'{kw_parts[0].title()} de {kw_parts[1].title()}'"
        elif len(kw_parts) == 1:
            ej_bueno = f"'{kw_parts[0].title()} en la región'"
        else:
            ej_bueno = "'Proyecto de terminal de transportes'"

        ej_malo = f"'{kw_parts[0].title()} {kw_parts[1].title()}' (sin preposición)" if len(kw_parts) >= 2 else "'Actividad corporativa', 'Gestión institucional'"

        instruccion_prep = (
            "  OBLIGATORIO: usa una preposición (de, del, para, sobre, en, por) "
            "entre los conceptos. NUNCA dos sustantivos pegados sin nexo.\n"
        ) if forzar_preposicion else ""

        instruccion_verbo = (
            "  PROHIBIDO: verbos conjugados ('presenta', 'anuncia', 'lanza', 'inaugura', etc.). "
            "Solo frases nominales (sustantivos + preposiciones).\n"
            "  NUNCA empieces con cargo ('Alcalde', 'Gobernador', 'Ministro', 'Director').\n"
        ) if prohibir_verbos else ""

        prompt = (
            f"Eres analista de reputación de '{self.marca}'. Genera UN subtema periodístico (3-5 palabras) "
            "como frase nominal sin verbo conjugado.\n\n"
            f"Títulos: {' | '.join(titulos[:5])}{ctx}\n"
            f"Keywords: {kw}\n\n"
            f"{instruccion_prep}{instruccion_verbo}"
            f"CORRECTO: {ej_bueno}, 'Tarifas de energía eléctrica'\n"
            f"INCORRECTO: {ej_malo}, 'Alcalde presenta plan'\n"
            "Tildes y ñ correctas. Describe el hecho vinculado con la marca; no respondas solo su nombre.\n"
            'JSON: {"subtema":"..."}'
        )
        try:
            resp = call_with_retries(
                openai.ChatCompletion.create,
                model=OPENAI_MODEL_CLASIFICACION,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=60,
                temperature=0.2,
                response_format={"type": "json_object"}
            )
            raw = json.loads(resp.choices[0].message.content).get("subtema", "Varios")
            et = limpiar_tema(raw)
            if not _frase_esta_completa(et):
                et = _recortar_frase_completa(et)
                if not _frase_esta_completa(et): return self._fallback(titulos)
            return et
        except:
            return self._fallback([])

    def _fallback(self, titulos):
        if not titulos: return "Cobertura de información relevante"
        texto_total = " ".join(str(t) for t in titulos[:5])
        norm_total = _normalizar_mencion(texto_total)

        # Respaldo genérico: identifica el tipo de hecho y su objeto, sin reglas por cliente.
        acciones = [
            (r"\b(lanzamiento|lanza|lanzo|presenta|presento|estrena|estreno)\b", "Lanzamiento"),
            (r"\b(anuncia|anuncio)\b", "Anuncio"),
            (r"\b(inaugura|inauguro|apertura|abre|abrio)\b", "Apertura"),
            (r"\b(firma|firmo|suscribe|suscribio|convenio|alianza)\b", "Convenio"),
            (r"\b(recibe|recibio|premio|reconocimiento)\b", "Reconocimiento"),
            (r"\b(investiga|investigacion|sancion|demanda)\b", "Investigación"),
        ]
        accion = next((nombre for patron, nombre in acciones if re.search(patron, norm_total)), None)
        palabras = []
        tokens_marca = set(_normalizar_mencion(" ".join([self.marca] + self.aliases)).split())
        excluir = tokens_marca | STOPWORDS_ES | {
            "universidad", "empresa", "compania", "corporacion", "fundacion", "institucion",
            "anuncio", "anuncia", "anuncio", "lanzamiento", "lanza", "presenta", "presencia",
            "invitado", "especial", "principal", "marca", "cliente",
        }
        for t in titulos[:5]:
            for w in string_norm_label(t).split():
                if len(w) >= 4 and w not in excluir: palabras.append(w)
        if palabras:
            top = [w for w, _ in Counter(palabras).most_common(3)]
            if accion:
                objeto = " ".join(top[:3])
                frase = _recortar_frase_completa(f"{accion} de {objeto}", MAX_PALABRAS_SUBTEMA)
                if _frase_esta_completa(frase) and not _es_nombre_o_fragmento_marca(frase, self.marca, self.aliases):
                    return capitalizar_etiqueta(frase)
            if len(top) >= 2:
                frase = f"{top[0]} de {top[1]}"
                if _frase_esta_completa(frase) and not _es_nombre_o_fragmento_marca(frase, self.marca, self.aliases):
                    return capitalizar_etiqueta(frase)
                return capitalizar_etiqueta(f"Asuntos de {top[0]} y {top[1]}")
            return capitalizar_etiqueta(f"Asuntos relacionados con {top[0]}")
        return "Cobertura de información relevante"

    def _consolidar_sinonimos_llm(self, subtemas_unicos):
        if len(subtemas_unicos) <= 1:
            return {s: s for s in subtemas_unicos}
            
        prompt = (
            "Eres un analista de datos. Tienes la siguiente lista de subtemas periodísticos:\n"
            f"{', '.join(subtemas_unicos)}\n\n"
            "Tu tarea es encontrar SUBTEMAS SINÓNIMOS que signifiquen exactamente lo mismo "
            "(aunque usen palabras ligeramente distintas) y unificarlos bajo el nombre más claro y representativo.\n"
            "REGLAS:\n"
            "1. NO fusiones temas que sean distintos (ej. 'Inversión en vías' y 'Mantenimiento de vías' son distintos).\n"
            "2. SÍ fusiona sinónimos (ej. 'Lanzamiento de plataforma web' y 'Estreno de portal digital').\n"
            "3. Devuelve un objeto JSON donde las claves sean los subtemas originales y el valor sea el subtema unificado.\n\n"
            'Ejemplo de salida:\n'
            '{"Tendencias de consumo de pollo": "Tendencias de consumo de pollo", "Hábitos de compra de aves": "Tendencias de consumo de pollo"}'
        )
        try:
            resp = call_with_retries(
                openai.ChatCompletion.create,
                model=OPENAI_MODEL_CLASIFICACION,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=1000,
                temperature=0.0,
                response_format={"type": "json_object"}
            )
            return json.loads(resp.choices[0].message.content)
        except:
            return {s: s for s in subtemas_unicos}

    def procesar_lote(self, col, pbar, res_puros, tit_puros):
        textos   = col.tolist()
        titulos  = tit_puros.tolist()
        resumenes = res_puros.tolist()
        n = len(textos)

        self._umbrales = _umbrales_adaptativos(n)
        u = self._umbrales
        st.caption(
            f"📐 Corpus: **{n}** noticias · Umbral subtema: **{u['subtema']}** · "
            f"Sim mínima: **{u['sim_minima_agrupacion']}**"
        )

        et = [texto_para_embedding(titulos[i], resumenes[i]) for i in range(n)]

        pbar.progress(0.05, "Fase 1 · Idénticas...")
        dsu = DSU(n)
        self._paso1(titulos, resumenes, dsu)
        
        pbar.progress(0.12, "Fase 2 · Títulos...")
        self._paso2(titulos, dsu)

        pbar.progress(0.18, "Embeddings...")
        ae = get_embeddings_batch(et)

        if u['usar_paso2b']:
            pbar.progress(0.15, "Fase 2b · Keywords raras (con validación semántica)...")
            self._paso2b_keywords(titulos, dsu, ae)

        pbar.progress(0.20, "Fase 3 · Clustering...")
        self._paso3(et, ae, dsu, pbar, 0.20)

        gf = dsu.grupos(n)
        ng = len(gf)
        pbar.progress(0.55, f"Fase 4 · Etiquetando {ng} grupos...")
        mapa = {}
        sg = sorted(gf.items(), key=lambda x: -len(x[1]))
        subtemas_aprobados = [] 
        textos_por_subtema_aprobado = defaultdict(list)

        def _generar_etiqueta_segura(idxs):
            # Sample the LLM prompt, but every member of this DSU group gets the same label.
            sample = idxs[:MAX_GRUPO_ETIQUETA]
            textos_grp = [textos[i] for i in sample]
            titulos_grp = [titulos[i] for i in sample]
            resumenes_grp = [resumenes[i] for i in sample]
            etiqueta = self._generar_etiqueta(
                textos_grp,
                titulos_grp,
                resumenes_grp,
                subtemas_existentes=subtemas_aprobados
            )
            if etiqueta in textos_por_subtema_aprobado:
                previos = textos_por_subtema_aprobado.get(etiqueta, [])
                if not _grupos_contenido_compatibles(
                    textos_grp,
                    previos,
                    etiqueta,
                    etiqueta,
                    min_sim=max(u['sim_minima_agrupacion'], 0.88),
                    min_overlap=0.24,
                ):
                    rechazada = etiqueta
                    etiqueta = self._generar_etiqueta(
                        textos_grp,
                        titulos_grp,
                        resumenes_grp,
                        subtemas_existentes=subtemas_aprobados,
                        evitar_etiqueta=rechazada
                    )
                    if etiqueta in textos_por_subtema_aprobado:
                        previos2 = textos_por_subtema_aprobado.get(etiqueta, [])
                        if not _grupos_contenido_compatibles(
                            textos_grp,
                            previos2,
                            etiqueta,
                            etiqueta,
                            min_sim=max(u['sim_minima_agrupacion'], 0.88),
                            min_overlap=0.24,
                        ):
                            etiqueta = capitalizar_etiqueta(self._fallback(titulos_grp))
            if etiqueta not in subtemas_aprobados:
                subtemas_aprobados.append(etiqueta)
            textos_por_subtema_aprobado[etiqueta].extend(textos_grp)
            return etiqueta

        for k, (lid, idxs) in enumerate(sg):
            if k % 10 == 0: pbar.progress(0.55 + 0.25 * (k / max(ng, 1)), f"Etiquetando {k + 1}/{ng}...")
            e = _generar_etiqueta_segura(idxs)
            for i in idxs: mapa[i] = e

        subtemas = [mapa.get(i, "Varios") for i in range(n)]

        pbar.progress(0.80, "Fase 4b · Coherencia (sin reasignar)...")
        # 0.35 cosine-to-label is not event membership. Jumping rows onto
        # another Subtema (or minting a new phrase) over-grouped and paraphrased.

        pbar.progress(0.86, "Fase 5 · Sin fusión cruzada de etiquetas...")
        # Skip corpus-wide dedup_labels / _fusionar_subtemas_semanticos.
        # Those glued distinct events that shared a 5-word paraphrase.

        pbar.progress(0.90, "Fase 6 · Consistencia...")
        subtemas = self._consistencia(subtemas, ae, pbar, u)

        indices_reclass = [i for i, s in enumerate(subtemas) if s == "_RECLASSIFICAR"]
        if indices_reclass:
            pbar.progress(0.93, f"Fase 6b · Reclasificando...")
            for i in indices_reclass:
                et_ind = self._generar_etiqueta([textos[i]], [titulos[i]], [resumenes[i]], subtemas_existentes=subtemas_aprobados)
                subtemas[i] = capitalizar_etiqueta(et_ind)
                if et_ind not in subtemas_aprobados: subtemas_aprobados.append(et_ind)

        pbar.progress(0.93, "Fase 7 · Completitud...")
        subtemas = self._validar_completitud_final(subtemas, textos, titulos, resumenes)

        pbar.progress(0.97, "Fase 8 · Sin dedup ni sinónimos cruzados...")
        subtemas = [capitalizar_etiqueta(s) for s in subtemas]
        nf = len(set(subtemas))
        pbar.progress(1.0, f"{nf} subtemas")
        st.info(f"Subtemas: **{nf}** · Grupos originales: **{ng}**")
        return subtemas

    def _validar_completitud_final(self, subtemas, textos, titulos, resumenes):
        por_subtema = defaultdict(list)
        for i, s in enumerate(subtemas): por_subtema[s].append(i)
        resultado = list(subtemas)
        for sub, idxs in por_subtema.items():
            if _frase_esta_completa(sub): continue
            recortada = _recortar_frase_completa(sub)
            if _frase_esta_completa(recortada) and len(recortada.split()) >= 2:
                for i in idxs: resultado[i] = capitalizar_etiqueta(recortada)
                continue
            tit_grp = [titulos[i] for i in idxs[:6]]
            res_grp = [resumenes[i] for i in idxs[:3]]
            nueva = _validar_etiqueta_completa(
                sub, titulos_grp=tit_grp, resumenes_grp=res_grp,
                marca=self.marca, aliases=self.aliases, fallback_fn=self._fallback
            )
            for i in idxs: resultado[i] = capitalizar_etiqueta(nueva)
        return resultado

    def _consistencia(self, subtemas, ae, pbar, umbrales=None):
        min_sub = umbrales.get('min_pertenencia_subtema', UMBRAL_MIN_PERTENENCIA_SUBTEMA)
        ps = defaultdict(list)
        for i, s in enumerate(subtemas): ps[s].append(i)
        r = list(subtemas)
        centroids = {}
        for sub, idxs in ps.items():
            vecs = [ae[i] for i in idxs if ae[i] is not None]
            if vecs: centroids[sub] = np.mean(vecs, axis=0)
        for sub in [s for s in centroids if len(ps[s]) >= 3]:
            idxs = ps[sub]
            if sub.lower() in ("sin tema", "varios") or len(idxs) < 3: continue
            vi = [(i, ae[i]) for i in idxs if ae[i] is not None]
            if len(vi) < 3: continue
            v_i, v_v = zip(*vi)
            M = np.array(v_v)
            sims = cosine_similarity(M, centroids[sub].reshape(1, -1)).flatten()
            thr = max(0.60, np.mean(sims) - 2 * np.std(sims))
            for k, (oi, sv) in enumerate(zip(v_i, sims)):
                if sv >= thr: continue
                bs, bsim = sub, sv
                emb = ae[oi]
                for os_, oc in centroids.items():
                    if os_ == sub: continue
                    s2 = cosine_similarity(np.array(emb).reshape(1, -1), oc.reshape(1, -1))[0][0]
                    if s2 > bsim and s2 > 0.75: bsim = s2; bs = os_
                if bs != sub: r[oi] = bs
                elif sv < min_sub: r[oi] = "_RECLASSIFICAR"
        return r

# ======================================
# TEMAS  
# ======================================
def _construir_representacion_grupo(subtema, textos_grupo, max_textos=30):
    palabras = []
    for t in textos_grupo[:max_textos]:
        for w in string_norm_label(str(t)).split():
            if len(w) > 3: palabras.append(w)
    kw_str = " ".join(w for w, _ in Counter(palabras).most_common(12))
    return f"{subtema}. {subtema}. {kw_str}"[:500]

def _validar_estructura_tema(tema: str) -> bool:
    if not tema or len(tema.split()) < 2: return False
    if len(tema.split()) > 5: return False
    if re.match(r'^[0-9]', tema): return False
    num_palabras = re.compile(
        r'^(uno|dos|tres|cuatro|cinco|seis|siete|ocho|nueve|diez|'
        r'once|doce|veinte|cien|varios|cada)', re.IGNORECASE
    )
    if num_palabras.match(tema): return False
    if _PATRON_TITULAR.match(tema): return False
    if _PATRON_ESTADO.search(tema): return False
    genericos = {
        "economia", "politica", "tecnologia", "seguridad", "justicia",
        "actualidad", "nacional", "internacional", "empresas", "sociedad",
        "negocios", "informacion", "noticias", "varios", "general",
    }
    if string_norm_label(tema) in genericos: return False
    return True

def _tema_es_igual_a_subtema(tema: str, subtemas_grupo: list) -> bool:
    if not tema or not subtemas_grupo: return False
    tn = string_norm_label(tema)
    for sub in subtemas_grupo:
        sn = string_norm_label(sub)
        if not tn or not sn: continue
        if SequenceMatcher(None, tn, sn).ratio() >= 0.80: return True
        if tn in sn or sn in tn: return True
    return False

def _generar_nombre_tema_llm(subtemas_grupo, textos_muestra, titulos_muestra, marca=""):
    subs_list = "\n".join(f"  · {s}" for s in subtemas_grupo[:8])
    palabras = []
    for t in titulos_muestra[:15]:
        for w in string_norm_label(str(t)).split():
            if len(w) > 3: palabras.append(w)
    kw = ", ".join(w for w, _ in Counter(palabras).most_common(6))
    tit_muestra = "\n".join(f"  · {t[:100]}" for t in list(dict.fromkeys(titulos_muestra))[:5])
    prompt = (
        f"Eres analista de reputación de la marca principal '{marca}'. "
        "Crea UN tema editorial preciso (2-5 palabras) que agrupe estos subtemas y describa el ámbito del hecho relacionado con la marca.\n\n"
        "SUBTEMAS:\n" + subs_list + "\n\nTÍTULOS DE REFERENCIA:\n" + tit_muestra +
        f"\n\nKEYWORDS: {kw}\n\n"
        "REGLAS ESTRICTAS:\n"
        "  1. Conserva el asunto común que diferencia este grupo; NO uses secciones vagas de una palabra.\n"
        "  2. Debe ser más general que los subtemas, pero no abstracto: nunca copies un titular ni repitas un subtema.\n"
        "  3. NUNCA incluyas números, cantidades ni nombres propios.\n"
        "  4. 2-5 palabras, sustantivo + complemento/adjetivo.\n"
        "  5. Tildes y ñ correctas.\n\n"
        "CORRECTO: 'Regulación financiera', 'Movilidad urbana', 'Infraestructura vial', 'Salud pública territorial'\n"
        "INCORRECTO: 'Economía', 'Política', 'Actualidad', 'Cinco congresistas con líos', 'Nuevo acuerdo'\n\n"
        'JSON: {"tema":"..."}'
    )
    try:
        resp = call_with_retries(
            openai.ChatCompletion.create,
            model=OPENAI_MODEL_CLASIFICACION,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=40,
            temperature=0.05,
            response_format={"type": "json_object"}
        )
        raw = json.loads(resp.choices[0].message.content).get("tema", "").strip().replace('"', '').replace('.', '')
        nombre = limpiar_tema(raw)
        if not _validar_estructura_tema(nombre): return None
        return nombre
    except:
        return None

def _regenerar_tema_diferente(subtemas_grupo, titulos_muestra, intento=0):
    subs_list = ", ".join(subtemas_grupo[:8])
    prompt = (
        f"Subtemas: {subs_list}\n\n"
        "Genera UNA categoría precisa (2-5 palabras), diferente a los subtemas. "
        "Conserva el asunto común; no respondas una sección vaga de una palabra como Economía, Política o Actualidad. "
        "Tildes y ñ correctas, terminar en sustantivo/adjetivo.\n"
        'JSON: {"tema":"..."}'
    )
    try:
        resp = call_with_retries(
            openai.ChatCompletion.create,
            model=OPENAI_MODEL_CLASIFICACION,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=50,
            temperature=0.2 + intento * 0.1,
            response_format={"type": "json_object"}
        )
        nombre = limpiar_tema(json.loads(resp.choices[0].message.content).get("tema", "").strip().replace('"', '').replace('.', ''))
        return nombre if _validar_estructura_tema(nombre) else None
    except:
        return None

def consolidar_temas(subtemas, textos, pbar, marca=""):
    n = len(textos)
    u = _umbrales_adaptativos(n)
    pbar.progress(0.05, "Preparando temas...")
    df = pd.DataFrame({'subtema': subtemas, 'texto': textos})
    us = list(df['subtema'].unique())
    if len(us) <= 1:
        pbar.progress(1.0, "Un tema")
        return [capitalizar_etiqueta(s) for s in subtemas]

    if n <= 5 and len(us) == n:
        pbar.progress(1.0, "Corpus pequeño: temas = subtemas")
        st.info(f"Temas: **{n}** (corpus pequeño — cada noticia tiene tema propio)")
        return [capitalizar_etiqueta(s) for s in subtemas]

    pbar.progress(0.10, "Representaciones...")
    textos_por_subtema = defaultdict(list)
    for i, sub in enumerate(subtemas): textos_por_subtema[sub].append(textos[i])
    repr_enriquecidas = [_construir_representacion_grupo(sub, textos_por_subtema[sub]) for sub in us]
    pbar.progress(0.20, "Embeddings contenido...")
    emb_repr = get_embeddings_batch(repr_enriquecidas)
    emb_labels = get_embeddings_batch(us)
    ae = get_embeddings_batch(textos)
    centroids_contenido = {}
    for sub in us:
        idxs = df.index[df['subtema'] == sub].tolist()[:50]
        vecs = [ae[i] for i in idxs if ae[i] is not None]
        if vecs: centroids_contenido[sub] = np.mean(vecs, axis=0)
    pbar.progress(0.35, "Similitudes...")
    vs = [s for s in us if s in centroids_contenido]
    if len(vs) < 2:
        pbar.progress(1.0, "Sin agrupación")
        return [capitalizar_etiqueta(s) for s in subtemas]
    idx_map = {s: i for i, s in enumerate(us)}
    M_content = np.array([centroids_contenido[s] for s in vs])
    sim_content = cosine_similarity(M_content)
    has_repr = all(emb_repr[idx_map[s]] is not None for s in vs)
    has_label = all(emb_labels[idx_map[s]] is not None for s in vs)
    if has_repr and has_label:
        sim_combined = (0.50 * sim_content + 0.35 * cosine_similarity(np.array([emb_repr[idx_map[s]] for s in vs])) + 0.15 * cosine_similarity(np.array([emb_labels[idx_map[s]] for s in vs])))
    elif has_repr:
        sim_combined = (0.60 * sim_content + 0.40 * cosine_similarity(np.array([emb_repr[idx_map[s]] for s in vs])))
    else:
        sim_combined = sim_content

    pbar.progress(0.45, "Clustering temas...")
    dist_matrix = np.clip(1 - sim_combined, 0, 2)
    np.fill_diagonal(dist_matrix, 0)
    umbral_tema = u['tema']
    num_temas_max = u['num_temas_max']
    linkage_temas = 'complete' if len(vs) <= 6 else 'average'
    cl = AgglomerativeClustering(
        n_clusters=None, distance_threshold=1 - umbral_tema,
        metric='precomputed', linkage=linkage_temas
    ).fit(dist_matrix)

    clusters = defaultdict(list)
    for i, lbl in enumerate(cl.labels_): clusters[lbl].append(vs[i])
    clusters_validados = {}
    next_cluster_id = 0
    for _, subs_cluster in clusters.items():
        if len(subs_cluster) <= 1:
            clusters_validados[next_cluster_id] = subs_cluster
            next_cluster_id += 1
            continue
        dsu_tema = DSU(len(subs_cluster))
        for i in range(len(subs_cluster)):
            for j in range(i + 1, len(subs_cluster)):
                sa, sb = subs_cluster[i], subs_cluster[j]
                if _grupos_contenido_compatibles(
                    textos_por_subtema.get(sa, []),
                    textos_por_subtema.get(sb, []),
                    sa,
                    sb,
                    min_sim=max(umbral_tema, 0.82),
                    min_overlap=0.16,
                ):
                    dsu_tema.union(i, j)
        for miembros in dsu_tema.grupos(len(subs_cluster)).values():
            clusters_validados[next_cluster_id] = [subs_cluster[i] for i in miembros]
            next_cluster_id += 1
    clusters = clusters_validados
    uc = [s for s in us if s not in vs]
    mt = {}
    tc = len(clusters)
    pbar.progress(0.50, f"Nombres {tc} temas...")
    for k, (cid, subtemas_cluster) in enumerate(clusters.items()):
        pbar.progress(0.50 + 0.35 * (k / max(tc, 1)), f"Tema {k + 1}/{tc}...")
        titulos_cluster = []
        textos_cluster = []
        for sub in subtemas_cluster:
            for idx in df.index[df['subtema'] == sub].tolist()[:10]:
                txt = str(textos[idx])
                partes = txt.split('. ')
                if partes: titulos_cluster.append(partes[0][:120])
                textos_cluster.append(txt[:200])
        if len(subtemas_cluster) == 1:
            sub_unico = subtemas_cluster[0]
            nombre = _generar_nombre_tema_llm(subtemas_cluster, textos_cluster, titulos_cluster, marca)
            if not nombre or _tema_es_igual_a_subtema(nombre, subtemas_cluster):
                nombre = _regenerar_tema_diferente(subtemas_cluster, titulos_cluster)
            if not nombre or _tema_es_igual_a_subtema(nombre, subtemas_cluster):
                p = sub_unico.split()
                nombre = _recortar_frase_completa(" ".join(p), max_palabras=3) if len(p) > 3 else sub_unico
                if _tema_es_igual_a_subtema(nombre, subtemas_cluster): nombre = sub_unico
        else:
            nombre = _generar_nombre_tema_llm(subtemas_cluster, textos_cluster, titulos_cluster, marca)
            if not nombre or _tema_es_igual_a_subtema(nombre, subtemas_cluster):
                nombre = _regenerar_tema_diferente(subtemas_cluster, titulos_cluster)
            if not nombre or _tema_es_igual_a_subtema(nombre, subtemas_cluster):
                nombre = _regenerar_tema_diferente(subtemas_cluster, titulos_cluster, intento=1)
            if not nombre or _tema_es_igual_a_subtema(nombre, subtemas_cluster):
                all_words = []
                for sub in subtemas_cluster:
                    for w in string_norm_label(sub).split():
                        if len(w) > 3: all_words.append(w)
                nombre = capitalizar_etiqueta(" ".join(w for w, _ in Counter(all_words).most_common(2))) if all_words else subtemas_cluster[0]
        if not _frase_esta_completa(nombre):
            nombre = _recortar_frase_completa(nombre, max_palabras=4)
            if not _frase_esta_completa(nombre):
                freq = Counter(subtemas)
                nombre = _recortar_frase_completa(max(subtemas_cluster, key=lambda s: freq.get(s, 0)), max_palabras=4)
        nombre = capitalizar_etiqueta(nombre)
        for sub in subtemas_cluster: mt[sub] = nombre
    for sub in uc: mt[sub] = capitalizar_etiqueta(sub)

    pbar.progress(0.87, "Validando pertenencia mínima a temas...")
    min_tema = u['min_pertenencia_tema']
    tf_inicial = [mt.get(sub, sub) for sub in subtemas]
    tema_agrupacion: Dict[str, list] = defaultdict(list)
    for i, tema in enumerate(tf_inicial):
        if ae[i] is not None: tema_agrupacion[tema].append(ae[i])
    tema_centroids: Dict[str, np.ndarray] = {
        t: np.mean(vecs, axis=0) for t, vecs in tema_agrupacion.items() if vecs
    }
    tf_validado: List[str] = []
    n_forzadas = 0
    for i, (sub, tema_asignado) in enumerate(zip(subtemas, tf_inicial)):
        emb = ae[i]
        if emb is not None and tema_asignado in tema_centroids:
            sim = cosine_similarity(np.array(emb).reshape(1, -1), tema_centroids[tema_asignado].reshape(1, -1))[0][0]
            if sim < min_tema:
                tf_validado.append(capitalizar_etiqueta(_recortar_frase_completa(sub, max_palabras=4)))
                n_forzadas += 1
                continue
        tf_validado.append(capitalizar_etiqueta(tema_asignado))
    if n_forzadas: st.caption(f"ℹ️ {n_forzadas} noticias con baja pertenencia al tema agrupado → tema propio asignado.")

    pbar.progress(0.88, "Dedup temas...")
    tf_validado = dedup_labels(tf_validado, u['dedup_label'])

    pbar.progress(0.90, "Fusionando temas solapados...")
    mapa_fusion_temas = _fusionar_temas_contenidos(tf_validado)
    if mapa_fusion_temas:
        tf_validado = [mapa_fusion_temas.get(t, t) for t in tf_validado]

    pbar.progress(0.92, "Validando tema ≠ subtema...")
    tf_validado = _post_validar_tema_vs_subtema(tf_validado, subtemas)
    pbar.progress(0.95, "Completitud...")
    tf_validado = [capitalizar_etiqueta(_recortar_frase_completa(t) if not _frase_esta_completa(t) else t) for t in tf_validado]
    tf_validado = _unificar_tema_por_subtema(tf_validado, subtemas)
    st.info(f"Temas: **{len(set(tf_validado))}** (de {len(set(subtemas))} subtemas) · Máx: {num_temas_max}")
    pbar.progress(1.0, "Temas listos")
    return tf_validado

def _fusionar_temas_contenidos(temas: List[str]) -> Dict[str, str]:
    unique = list(dict.fromkeys(temas))
    if len(unique) < 2: return {}
    normed = {t: string_norm_label(t) for t in unique}
    mapa: Dict[str, str] = {}
    for i, ta in enumerate(unique):
        for tb in unique[i + 1:]:
            na, nb = normed[ta], normed[tb]
            if not na or not nb: continue
            if na == nb or SequenceMatcher(None, na, nb).ratio() >= 0.92:
                canon = tb if len(tb) >= len(ta) else ta
                reemplazar = ta if canon == tb else tb
                mapa[reemplazar] = canon
    umbral_relajado = 0.88
    candidatos = [(t, normed[t]) for t in unique if len(t.split()) <= 3 and t not in mapa]
    if len(candidatos) >= 2:
        textos_c = [t for t, _ in candidatos]
        embs = get_embeddings_batch(textos_c)
        validos = [(textos_c[i], embs[i]) for i in range(len(textos_c)) if embs[i] is not None]
        if len(validos) >= 2:
            etqs, vecs = zip(*validos)
            sim = cosine_similarity(np.array(vecs))
            for i in range(len(etqs)):
                for j in range(i + 1, len(etqs)):
                    if sim[i][j] >= umbral_relajado:
                        ta, tb = etqs[i], etqs[j]
                        if ta in mapa or tb in mapa: continue
                        if _etiquetas_compatibles(ta, tb, min_overlap=0.60):
                            freq = Counter(temas)
                            canon = ta if freq.get(ta, 0) >= freq.get(tb, 0) else tb
                            reemplazar = tb if canon == ta else ta
                            mapa[reemplazar] = canon
    return mapa

def _post_validar_tema_vs_subtema(temas, subtemas):
    tema_a_subtemas = defaultdict(set)
    for t, s in zip(temas, subtemas): tema_a_subtemas[t].add(s)
    reemplazos = {}
    for tema, subs in tema_a_subtemas.items():
        if len(subs) == 1:
            sub_unico = list(subs)[0]
            tn = string_norm_label(tema)
            sn = string_norm_label(sub_unico)
            if tn and sn and SequenceMatcher(None, tn, sn).ratio() >= 0.80:
                nuevo = _regenerar_tema_diferente([sub_unico], [])
                if nuevo and not _tema_es_igual_a_subtema(nuevo, [sub_unico]) and _frase_esta_completa(nuevo):
                    reemplazos[tema] = capitalizar_etiqueta(nuevo)
    return [reemplazos.get(t, t) for t in temas] if reemplazos else temas

def _unificar_tema_por_subtema(temas, subtemas):
    """Un mismo Subtema (sin importar mayúsculas) debe tener un único Tema."""
    vacios = {"", "nan", "n/a", "-", "sin tema", "varios"}
    sub_to_temas = defaultdict(list)
    for t, s in zip(temas, subtemas):
        k = string_norm_label(s)
        if not k or k in vacios:
            continue
        sub_to_temas[k].append(t)
    sub_to_best = {}
    for k, tema_list in sub_to_temas.items():
        validos = [t for t in tema_list if str(t).strip().lower() not in vacios]
        if validos:
            sub_to_best[k] = Counter(validos).most_common(1)[0][0]
    out = []
    for t, s in zip(temas, subtemas):
        k = string_norm_label(s)
        out.append(sub_to_best[k] if k in sub_to_best else t)
    return out

# ======================================
# Duplicados y Excel (Reglas Nuevas)
# ======================================
def _normalizar_url(url: str) -> str:
    if not url: return ""
    url = url.strip().lower()
    url = re.sub(r'^https?://', '', url)
    url = re.sub(r'^www\.', '', url)
    url = url.rstrip('/')
    return url

def detectar_duplicados_avanzado(rows, km):
    processed = deepcopy(rows)
    seen_url, seen_bcast = {}, {}
    seen_streaming: Dict[tuple, int] = {}
    tb = defaultdict(list)

    for i, row in enumerate(processed):
        if row.get("is_duplicate"): continue

        tipo    = normalizar_tipo_medio(str(row.get(km["tipodemedio"], "")))
        mencion = norm_key(row.get(km["menciones"], ""))
        medio   = norm_key(row.get(km["medio"], ""))

        streaming_url_raw = row.get(km["link_streaming"])
        if isinstance(streaming_url_raw, dict):
            streaming_url_raw = streaming_url_raw.get("url")
            
        if streaming_url_raw and mencion:
            streaming_url_norm = _normalizar_url(str(streaming_url_raw))
            if streaming_url_norm:
                sk = (streaming_url_norm, mencion)
                if sk in seen_streaming:
                    row["is_duplicate"] = True
                    row[km["idduplicada"]] = processed[seen_streaming[sk]].get(km["idnoticia"], "")
                    continue
                seen_streaming[sk] = i

        if tipo == "Internet":
            li = row.get(km["link_nota"])
            url = li.get("url") if isinstance(li, dict) else li
            if url and mencion:
                url_norm = _normalizar_url(str(url))
                k = (url_norm, mencion)
                if k in seen_url:
                    row["is_duplicate"] = True
                    row[km["idduplicada"]] = processed[seen_url[k]].get(km["idnoticia"], "")
                    continue
                seen_url[k] = i
            if medio and mencion:
                tb[(medio, mencion)].append(i)

        elif tipo in ("Radio", "Televisión"):
            hora = str(row.get(km["hora"], "")).strip()
            if mencion and medio and hora:
                k = (mencion, medio, hora)
                if k in seen_bcast:
                    row["is_duplicate"] = True
                    row[km["idduplicada"]] = processed[seen_bcast[k]].get(km["idnoticia"], "")
                else:
                    seen_bcast[k] = i

    for idxs in tb.values():
        if len(idxs) < 2: continue
        for i in range(len(idxs)):
            for j in range(i + 1, len(idxs)):
                a, b = idxs[i], idxs[j]
                if processed[a].get("is_duplicate") or processed[b].get("is_duplicate"): continue
                ta  = normalize_title_for_comparison(processed[a].get(km["titulo"]))
                tb_ = normalize_title_for_comparison(processed[b].get(km["titulo"]))
                if ta and tb_ and SequenceMatcher(None, ta, tb_).ratio() >= SIMILARITY_THRESHOLD_TITULOS:
                    if len(ta) < len(tb_):
                        processed[a]["is_duplicate"] = True
                        processed[a][km["idduplicada"]]  = processed[b].get(km["idnoticia"], "")
                    else:
                        processed[b]["is_duplicate"] = True
                        processed[b][km["idduplicada"]]  = processed[a].get(km["idnoticia"], "")

    return processed

def read_and_normalize_dossier(sheet, region_map, internet_map):
    headers = [cell.value for cell in sheet[1] if cell.value is not None]
    rows = []
    for row in sheet.iter_rows(min_row=2):
        if all(c.value is None for c in row):
            continue
        row_data = {}
        for i, h in enumerate(headers):
            if i < len(row):
                cell = row[i]
                val = cell.value
                url = cell.hyperlink.target if (cell.hyperlink and cell.hyperlink.target) else None
                if url:
                    row_data[h] = {"value": val or "Link", "url": url}
                else:
                    row_data[h] = val
        rows.append(row_data)

    df = pd.DataFrame(rows)

    tipo_medio_map = {
        'online': 'Internet', 'internet': 'Internet',
        'diario': 'Prensa',
        'am': 'Radio', 'fm': 'Radio',
        'aire': 'Televisión', 'cable': 'Televisión',
        'revista': 'Revistas', 'revistas': 'Revistas',
    }
    
    if 'Tipo de Medio' in df.columns:
        df['Tipo de Medio'] = (
            df['Tipo de Medio'].astype(str).str.lower().str.strip()
            .map(tipo_medio_map)
            .fillna(df['Tipo de Medio'].astype(str).str.strip())
        )
    else:
        df['Tipo de Medio'] = 'Otro'

    is_av = df['Tipo de Medio'].isin(['Radio', 'Televisión'])
    is_grafica = df['Tipo de Medio'].isin(['Prensa', 'Internet', 'Revistas'])
    is_internet = df['Tipo de Medio'] == 'Internet'

    if 'Medio' in df.columns:
        raw_medios_clean = df['Medio'].astype(str).str.lower().str.strip()
        df['Región'] = raw_medios_clean.map(region_map).fillna("N/A")
    else:
        df['Medio'] = 'N/A'
        df['Región'] = 'N/A'

    if 'Medio' in df.columns:
        df.loc[is_internet, 'Medio'] = (
            df.loc[is_internet, 'Medio']
            .astype(str).str.lower().str.strip()
            .map(internet_map)
            .fillna(df.loc[is_internet, 'Medio'])
        )

    df['ID Noticia'] = df.get('NoticiaId', df.get('ID Noticia', pd.Series(dtype=str)))
    df['Fecha'] = pd.to_datetime(df.get('Fecha', pd.Series(dtype=str)), dayfirst=True, errors='coerce').dt.normalize()
    df['Hora'] = df.get('Hora', pd.Series(dtype=str))
    df['Sección - Programa'] = df.get('Sección - Programa', pd.Series(dtype=str)).astype(str).apply(clean_text)
    
    titulo_col = 'Título' if 'Título' in df.columns else 'Titulo'
    df['Título'] = df.get(titulo_col, pd.Series(dtype=str)).astype(str).apply(clean_text)
    df['Autor - Conductor'] = df.get('Autor - Conductor', pd.Series(dtype=str)).astype(str).apply(clean_text)
    df['Nro. Pagina'] = df.get('Nro. Pagina', pd.Series(dtype=str))
    
    dim_col = 'Dimensioncm2' if 'Dimensioncm2' in df.columns else 'Dimensión'
    df['Dimensión'] = df.get(dim_col, pd.Series(dtype=str))
    df['Duración - Nro. Caracteres'] = df.get('Duración - Nro. Caracteres', pd.Series(dtype=str))

    df.loc[is_av, 'Dimensión'] = df.loc[is_av, 'Duración - Nro. Caracteres']
    df.loc[is_av, 'Duración - Nro. Caracteres'] = 0

    cpe_av = df.get('CPE', pd.Series([np.nan] * len(df)))
    cpe_grafica = df.get('Valor de Nota', pd.Series([np.nan] * len(df)))
    df['CPE'] = np.where(is_av, cpe_av, np.where(is_grafica, cpe_grafica, np.nan))

    df['Tier'] = df.get('Tier', pd.Series(dtype=str))
    df['Audiencia'] = df.get('Audiencia', pd.Series(dtype=str))
    df['Tono'] = df.get('Tono', pd.Series(dtype=str)).astype(str).apply(clean_text)
    df['Tema'] = df.get('Tematica', df.get('Tema', pd.Series(dtype=str))).astype(str).apply(clean_text)
    df['Temas Generales - Tema'] = df.get('Temas Generales - Tema', pd.Series(dtype=str)).astype(str).apply(clean_text)

    cuerpo_col = 'CuerpoEs' if 'CuerpoEs' in df.columns else 'Resumen - Aclaracion'
    cuerpo_cleaned = df.get(cuerpo_col, pd.Series([''] * len(df))).astype(str).apply(clean_cuerpo)

    def fmt_grafica(text):
        if not isinstance(text, str) or not text.strip():
            return text
        parrafos = [p.strip() for p in text.split('\n') if p.strip()]
        return '\n\n'.join(parrafos) if len(parrafos) > 1 else text

    df['Resumen - Aclaracion'] = np.where(is_av, cuerpo_cleaned, cuerpo_cleaned.apply(fmt_grafica))

    # ── ADICIÓN: columna con el CuerpoEs COMPLETO, sin truncar ──────────────
    # Se guarda tal cual queda cuerpo_cleaned (HTML limpio, <br> -> saltos de línea),
    # SIN pasar por corregir_texto() (que es lo que recorta/añade "..." al final).
    df['Cuerpo Completo'] = cuerpo_cleaned

    url_nota_av = df.get('URL Nota AV', df.get('Link Nota AV', pd.Series([''] * len(df))))
    url_streaming = df.get('URL (Streaming - Imagen)', pd.Series([''] * len(df)))
    
    link_nota_final = []
    for val_av, val_str, is_av_row in zip(url_nota_av, url_streaming, is_av):
        if is_av_row:
            if isinstance(val_av, dict):
                url_t = val_av.get("url", "")
                link_nota_final.append({"value": "Link", "url": url_t.replace(".com.ar", ".com.co") if url_t else None})
            else:
                url_t = str(val_av or "")
                link_nota_final.append({"value": "Link", "url": url_t.replace(".com.ar", ".com.co") if url_t else None})
        else:
            if isinstance(val_str, dict):
                link_nota_final.append(val_str)
            else:
                link_nota_final.append({"value": "Link", "url": val_str if val_str else None})
                
    df['Link Nota'] = link_nota_final

    url_nota_raw = df.get('URL Nota', pd.Series([''] * len(df)))
    link_stream_final = []
    for val_url, is_int in zip(url_nota_raw, is_internet):
        if is_int:
            if isinstance(val_url, dict):
                link_stream_final.append(val_url)
            else:
                link_stream_final.append({"value": "Link", "url": val_url if val_url else None})
        else:
            link_stream_final.append(None)
            
    df['Link (Streaming - Imagen)'] = link_stream_final

    menciones_av = df.get('Menciones - Empresa', pd.Series([''] * len(df))).fillna('').astype(str).apply(clean_text)
    menciones_grafica = df.get('Empresa rel.', pd.Series([''] * len(df))).fillna('').astype(str).apply(clean_text)
    df['Menciones - Empresa'] = np.where(is_av, menciones_av, np.where(is_grafica, menciones_grafica, menciones_av))

    return df

def generate_output_excel(rows, km):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado"
    ORDER = [
        "ID Noticia", "Fecha", "Hora", "Medio", "Tipo de Medio",
        "Sección - Programa", "Región", "Título", "Autor - Conductor",
        "Nro. Pagina", "Dimensión", "Duración - Nro. Caracteres",
        "CPE", "Tier", "Audiencia", "Tono", "Tono IA", "Tema", "Subtema", "Grupo noticia",
        "Link Nota", "Resumen - Aclaracion", "Link (Streaming - Imagen)", "Menciones - Empresa",
        "ID duplicada",
        "Cuerpo Completo"   # ── ADICIÓN: columna final con el CuerpoEs completo, sin truncar ──
    ]
    NUM = {"ID Noticia", "Nro. Pagina", "Dimensión", "Duración - Nro. Caracteres", "CPE", "Tier", "Audiencia"}
    ORDER += ["Contexto analizado", "Coincidencia marca", "Origen coincidencia"]
    ws.append(ORDER)
    
    font_hyperlink = Font(color="000000", underline=None)
    align_left = Alignment(horizontal='left')
    font_header = Font(bold=True)
    
    for i, col_name in enumerate(ORDER, start=1):
        cell = ws.cell(row=1, column=i)
        cell.font = font_header

    col_idx_map = {name: ORDER.index(name) + 1 for name in ORDER}
        
    for row in rows:
        ctx, match, origin = _brand_audit(row.get(km.get("titulo"), ""), row.get(km.get("resumen"), ""), st.session_state.get("brand_name", ""), st.session_state.get("brand_aliases", []))
        row["Contexto analizado"], row["Coincidencia marca"], row["Origen coincidencia"] = ctx, match, origin
        tk = km.get("titulo")
        if tk and tk in row: row[tk] = clean_title_for_output(row.get(tk))
        rk = km.get("resumen")
        if rk and rk in row: row[rk] = corregir_texto(row.get(rk))
        
        out, links = [], {}
        for ci, h in enumerate(ORDER, start=1):
            dk = km.get(norm_key(h), norm_key(h))
            val = row.get(h)
            cv = None
            
            if h == 'Fecha' and pd.notna(val):
                if isinstance(val, pd.Timestamp):
                    cv = val.to_pydatetime()
                elif isinstance(val, (datetime.datetime, datetime.date)):
                    cv = val
                else:
                    cv = str(val) if val is not None else None
            elif h in NUM:
                cv = parse_numeric(val)
            elif isinstance(val, dict) and "url" in val:
                cv = val.get("value", "Link")
                if val.get("url"): links[ci] = val["url"]
            elif val is not None:
                if isinstance(val, str) and val.startswith("http"):
                    cv = "Link"
                    links[ci] = val
                else:
                    cv = str(val)
            out.append(cv)
        ws.append(out)
        
        current_row = ws.max_row
        for ci, url in links.items():
            cell = ws.cell(row=current_row, column=ci)
            cell.hyperlink = url
            cell.font = font_hyperlink
            cell.alignment = align_left
            
        date_col_idx = ORDER.index("Fecha") + 1
        date_cell = ws.cell(row=current_row, column=date_col_idx)
        if isinstance(date_cell.value, (datetime.datetime, datetime.date)):
            date_cell.number_format = 'DD/MM/YYYY'
            
        cols_millares = ["Nro. Pagina", "Dimensión", "Duración - Nro. Caracteres", "Tier", "Audiencia"]
        for col_name in cols_millares:
            col_idx = col_idx_map[col_name]
            cell = ws.cell(row=current_row, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'

        cpe_idx = col_idx_map["CPE"]
        cpe_cell = ws.cell(row=current_row, column=cpe_idx)
        if isinstance(cpe_cell.value, (int, float)):
            cpe_cell.number_format = '$#,##0'
            
    for i, col_name in enumerate(ORDER, start=1):
        letter = ws.cell(row=1, column=i).column_letter
        if col_name in ['Título', 'Resumen - Aclaracion', 'Cuerpo Completo']:
            ws.column_dimensions[letter].width = 50
        elif col_name in ['Link Nota', 'Link (Streaming - Imagen)']:
            ws.column_dimensions[letter].width = 15
        else:
            ws.column_dimensions[letter].width = 20
            
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ======================================
# Proceso principal
# ======================================
async def run_full_process_async(df_file, bn, ba, tpkl, epkl, mode, xlsx_bytes=None, cliente="", voceros="", enable_scraping=False):
    st.session_state.update({'tokens_input': 0, 'tokens_output': 0, 'tokens_embedding': 0})
    get_embedding_cache().clear()
    t0 = time.time()
    
    if "API" in mode:
        try:
            openai.api_key=st.secrets["OPENAI_API_KEY"]
            openai.aiosession.set(None)
        except:
            st.error("OPENAI_API_KEY no encontrado.")
            st.stop()
            
    with st.status("Paso 1 · Carga de Configuración y Dossier", expanded=True) as s:
        region_map, internet_map = load_config_from_sheets()

        wb_in = load_workbook(df_file, data_only=True)
        df_normalized = read_and_normalize_dossier(wb_in.active, region_map, internet_map)

        medios_sin_region = sorted(set(
            df_normalized.loc[df_normalized['Región'] == 'N/A', 'Medio']
            .astype(str).str.strip()
        ) - {'', 'nan', 'None'})
        if medios_sin_region:
            st.session_state["medios_sin_mapear"] = medios_sin_region
        
        rows_expanded = []
        for idx, row_series in df_normalized.iterrows():
            menciones = [m.strip() for m in str(row_series['Menciones - Empresa']).split(';') if m.strip()]
            if not menciones:
                row_dict = row_series.to_dict()
                row_dict['Menciones - Empresa'] = ""
                row_dict['original_index'] = idx
                row_dict['expanded_index'] = len(rows_expanded)
                row_dict['is_duplicate'] = False
                rows_expanded.append(row_dict)
            else:
                for m in menciones:
                    row_dict = row_series.to_dict()
                    row_dict['Menciones - Empresa'] = m
                    row_dict['original_index'] = idx
                    row_dict['expanded_index'] = len(rows_expanded)
                    row_dict['is_duplicate'] = False
                    rows_expanded.append(row_dict)

        km = {
            "idnoticia": "ID Noticia",
            "fecha": "Fecha",
            "hora": "Hora",
            "medio": "Medio",
            "tipodemedio": "Tipo de Medio",
            "seccion_programa": "Sección - Programa",
            "region": "Región",
            "titulo": "Título",
            "autor_conductor": "Autor - Conductor",
            "nro_pagina": "Nro. Pagina",
            "dimension": "Dimensión",
            "duracion_caracteres": "Duración - Nro. Caracteres",
            "cpe": "CPE",
            "tier": "Tier",
            "audiencia": "Audiencia",
            "tono": "Tono",
            "tonoiai": "Tono IA",
            "tema": "Tema",
            "subtema": "Subtema",
            "link_nota": "Link Nota",
            "resumen": "Resumen - Aclaracion",
            "link_streaming": "Link (Streaming - Imagen)",
            "menciones": "Menciones - Empresa",
            "idduplicada": "ID duplicada"
        }
        
        rows = detectar_duplicados_avanzado(rows_expanded, km)
        for row in rows:
            if row["is_duplicate"]:
                row["Tono IA"] = "Duplicada"
                row["Tema"] = "-"
                row["Subtema"] = "-"
                
        s.update(label="✓ Paso 1 completado", state="complete")
        
    with st.status("Paso 2 · Normalización", expanded=True) as s:
        s.update(label="✓ Paso 2 · Mapeos y normalizaciones aplicados", state="complete")
        
    gc.collect()
    ta = [r for r in rows if not r.get("is_duplicate")]
    
    if ta:
        df = pd.DataFrame(ta)
        df["_txt"] = df.apply(
            lambda r: texto_para_embedding(str(r.get(km["titulo"], "")), str(r.get(km["resumen"], ""))),
            axis=1
        )
        with st.status("Embeddings...", expanded=True) as s:
            _ = get_embeddings_batch(df["_txt"].tolist())
            s.update(label=f"✓ {get_embedding_cache().stats()}", state="complete")
            
        with st.status("Paso 3 · Tono (Reputación)", expanded=True) as s:
            pb = st.progress(0)
            if ("PKL" in mode or tpkl) and tpkl:
                res = analizar_tono_con_pkl(
                    df["_txt"].tolist(), tpkl,
                    titulos=df[km["titulo"]], resumenes=df[km["resumen"]],
                    marca=bn, aliases=ba,
                )
                if res is None: st.stop()
            elif "API" in mode or "Híbrido" in mode:
                res = await ClasificadorTono(bn, ba).procesar_lote_async(
                    df["_txt"], pb, df[km["resumen"]], df[km["titulo"]]
                )
            else:
                res = [{"tono": "N/A"}] * len(ta)
            df[km["tonoiai"]] = [r["tono"] for r in res]
            s.update(label="✓ Paso 3 · Tono (Reputación)", state="complete")
            
        with st.status("Paso 4 · Clasificación", expanded=True) as s:
            pb = st.progress(0)
            if "Solo Modelos PKL" in mode:
                subtemas = ["N/A"] * len(ta)
                temas    = ["N/A"] * len(ta)
            else:
                subtemas = ClasificadorSubtema(bn, ba).procesar_lote(
                    df["_txt"], pb, df[km["resumen"]], df[km["titulo"]]
                )
                temas = consolidar_temas(subtemas, df["_txt"].tolist(), pb, bn)
            df[km["subtema"]] = subtemas
            if epkl:
                tp = analizar_temas_con_pkl(df["_txt"].tolist(), epkl)
                if tp: df[km["tema"]] = tp
            else:
                df[km["tema"]] = temas
            df[km["tema"]] = _unificar_tema_por_subtema(df[km["tema"]].tolist(), df[km["subtema"]].tolist())
            df = aplicar_consistencia_grupos(df, km["titulo"], km["resumen"],
                                             km["tonoiai"], km["tema"], km["subtema"])
            s.update(label="✓ Paso 4 · Clasificación", state="complete")
            
        rm2 = df.set_index("expanded_index").to_dict("index")
        for idx, row in enumerate(rows):
            if not row.get("is_duplicate"):
                row.update(rm2.get(row.get("expanded_index"), {}))
                
    gc.collect()
    ci = (st.session_state['tokens_input']     / 1e6) * PRICE_INPUT_1M
    co = (st.session_state['tokens_output']    / 1e6) * PRICE_OUTPUT_1M
    ce = (st.session_state['tokens_embedding'] / 1e6) * PRICE_EMBEDDING_1M
    
    st.session_state["brand_name"] = bn
    st.session_state["brand_aliases"] = ba
    with st.status("Paso 5 · Informe", expanded=True) as s:
        st.session_state["output_data"]     = generate_output_excel(rows, km)
        st.session_state["output_filename"] = f"Informe_IA_{bn.replace(' ', '_')}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        st.session_state["processing_complete"] = True
        st.session_state.update({
            "brand_name": bn, "brand_aliases": ba,
            "total_rows": len(rows), "unique_rows": len(ta), "duplicates": len(rows) - len(ta),
            "process_duration": f"{time.time() - t0:.0f}s",
            "process_cost": f"${ci + co + ce:.4f} USD",
            "cache_stats": get_embedding_cache().stats()
        })
        s.update(label=f"✓ Completado · {get_embedding_cache().stats()}", state="complete")

async def run_quick_async(df, tc, sc, bn, al):
    st.session_state.update({'tokens_input': 0, 'tokens_output': 0, 'tokens_embedding': 0})
    get_embedding_cache().clear()
    df['_txt'] = df.apply(lambda r: texto_para_embedding(str(r.get(tc, "")), str(r.get(sc, ""))), axis=1)
    with st.status("Embeddings...", expanded=True) as s:
        _ = get_embeddings_batch(df['_txt'].tolist())
        s.update(label=f"✓ {get_embedding_cache().stats()}", state="complete")
    with st.status("Tono", expanded=True) as s:
        pb = st.progress(0)
        res = await ClasificadorTono(bn, al).procesar_lote_async(df["_txt"], pb, df[sc].fillna(''), df[tc].fillna(''))
        df['Tono IA'] = [r["tono"] for r in res]
        audits = [_brand_audit(r.get(tc, ''), r.get(sc, ''), bn, al) for _, r in df.iterrows()]
        df['Contexto analizado'], df['Coincidencia marca'], df['Origen coincidencia'] = zip(*audits)
        s.update(label="✓ Tono", state="complete")
    with st.status("Clasificación", expanded=True) as s:
        pb = st.progress(0)
        subtemas = ClasificadorSubtema(bn, al).procesar_lote(df["_txt"], pb, df[sc].fillna(''), df[tc].fillna(''))
        df['Subtema'] = subtemas
        temas = consolidar_temas(subtemas, df["_txt"].tolist(), pb, bn)
        df['Tema'] = _unificar_tema_por_subtema(temas, subtemas)
        df = aplicar_consistencia_grupos(df, tc, sc)
        s.update(label="✓ Clasificación", state="complete")
    df.drop(columns=['_txt'], inplace=True)
    ci = (st.session_state['tokens_input']     / 1e6) * PRICE_INPUT_1M
    co = (st.session_state['tokens_output']    / 1e6) * PRICE_OUTPUT_1M
    ce = (st.session_state['tokens_embedding'] / 1e6) * PRICE_EMBEDDING_1M
    st.session_state['quick_cost'] = f"${ci + co + ce:.4f} USD"
    return df

def gen_quick_excel(df, original_bytes=None):
    if original_bytes:
        wb = load_workbook(io.BytesIO(original_bytes))
        ws = wb.active
        start = ws.max_column + 1
        for offset, col in enumerate([c for c in df.columns if c not in list(ws.values)[0]], start):
            ws.cell(1, offset, col)
            for i, value in enumerate(df[col].tolist(), 2): ws.cell(i, offset, value)
        out = io.BytesIO(); wb.save(out); return out.getvalue()
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        df.to_excel(w, index=False, sheet_name='Analisis')
    return buf.getvalue()

def render_quick_tab():
    st.markdown('<div class="sec-label">Análisis rápido</div>', unsafe_allow_html=True)
    if 'quick_result' in st.session_state:
        st.markdown(
            '<div class="success-banner"><div class="success-icon">✓</div>'
            '<div><div class="success-title">Completado</div>'
            '<div class="success-sub">Listo para descargar</div></div></div>',
            unsafe_allow_html=True
        )
        st.metric("Costo", st.session_state.get('quick_cost', "$0.00"))
        st.dataframe(st.session_state.quick_result.head(10), use_container_width=True)
        st.download_button(
            "Descargar",
            data=gen_quick_excel(st.session_state.quick_result, st.session_state.get('quick_bytes')),
            file_name="Analisis_Rapido_IA.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
        if st.button("Nuevo análisis"):
            for k in ('quick_result', 'quick_df', 'quick_name', 'quick_cost'):
                if k in st.session_state:
                    del st.session_state[k]
            st.rerun()
        return
    if 'quick_df' not in st.session_state:
        st.markdown("Sube un Excel con columnas de título y resumen.")
        f = st.file_uploader("Excel", type=["xlsx"], label_visibility="collapsed", key="qu")
        if f:
            try:
                st.session_state.quick_bytes = f.getvalue()
                st.session_state.quick_df   = pd.read_excel(io.BytesIO(st.session_state.quick_bytes))
                st.session_state.quick_name = f.name
                st.rerun()
            except Exception as e:
                st.error(f"Error: {e}")
    else:
        st.success(f"**{st.session_state.quick_name}** cargado")
        with st.form("qf"):
            cols = st.session_state.quick_df.columns.tolist()
            c1, c2 = st.columns(2)
            tc = c1.selectbox("Col. título", cols, _default_text_column_index(cols, ['Título', 'Titulo', 'Titular', 'Headline'], 0))
            sc = c2.selectbox("Col. resumen", cols, _default_text_column_index(cols, ['CuerpoEs', 'Resumen - Aclaración', 'Resumen - Aclaracion', 'Resumen', 'Cuerpo', 'Descripción', 'Descripcion'], 1))
            bn  = st.text_input("Marca",       placeholder="Ej: Bancolombia")
            bat = st.text_input("Alias (;)",   placeholder="Ej: Grupo Bancolombia;Ban")
            if st.form_submit_button("Analizar", use_container_width=True, type="primary"):
                if not bn:
                    st.error("Indica la marca.")
                else:
                    try:
                        openai.api_key = st.secrets["OPENAI_API_KEY"]
                        openai.aiosession.set(None)
                    except:
                        st.error("OPENAI_API_KEY no encontrada.")
                        st.stop()
                    al = [a.strip() for a in bat.split(";") if a.strip()]
                    with st.spinner("Procesando..."):
                        st.session_state.quick_result = asyncio.run(
                            run_quick_async(st.session_state.quick_df.copy(), tc, sc, bn, al)
                        )
                    st.rerun()
        if st.button("Otro archivo"):
            for k in ('quick_df', 'quick_name', 'quick_result', 'quick_cost'):
                if k in st.session_state:
                    del st.session_state[k]
            st.rerun()


# ======================================
# EXCEL PERSONALIZADO (Mantiene formato original + 3 columnas al final)
# ======================================
async def run_custom_excel_async(file_bytes, tc, sc, bn, al, mode="API de OpenAI", tpkl=None, epkl=None):
    st.session_state.update({'tokens_input': 0, 'tokens_output': 0, 'tokens_embedding': 0})
    get_embedding_cache().clear()
    t0 = time.time()

    # Cargar archivo usando openpyxl para conservar estilos y formato original
    buf_in = io.BytesIO(file_bytes)
    wb = load_workbook(buf_in)
    ws = wb.active

    # Cargar DataFrame solo para extraer textos e índices
    buf_in.seek(0)
    df = pd.read_excel(buf_in)

    df['_txt'] = df.apply(
        lambda r: texto_para_embedding(str(r.get(tc, "")), str(r.get(sc, ""))),
        axis=1
    )

    with st.status("Paso 1 · Generando Embeddings...", expanded=True) as s:
        _ = get_embeddings_batch(df['_txt'].tolist())
        s.update(label=f"✓ Embeddings listos · {get_embedding_cache().stats()}", state="complete")

    # --- PASO 2: TONO ---
    with st.status("Paso 2 · Evaluando Tono (Reputación)...", expanded=True) as s:
        pb = st.progress(0)
        if tpkl:
            # PKL de tono: predecir sobre la mención a Marca principal / Alias, no el artículo entero.
            res = analizar_tono_con_pkl(
                df["_txt"].tolist(), tpkl,
                titulos=df[tc].fillna(""), resumenes=df[sc].fillna(""),
                marca=bn, aliases=al,
            )
            if res is None: st.stop()
            tonos = [r["tono"] for r in res]
        elif "API" in mode or "Híbrido" in mode:
            res = await ClasificadorTono(bn, al).procesar_lote_async(
                df["_txt"], pb, df[sc].fillna(''), df[tc].fillna('')
            )
            tonos = [r["tono"] for r in res]
        else:
            tonos = ["N/A"] * len(df)
        df['Tono IA'] = tonos
        audits = [_brand_audit(r.get(tc, ''), r.get(sc, ''), bn, al) for _, r in df.iterrows()]
        df['Contexto analizado'], df['Coincidencia marca'], df['Origen coincidencia'] = zip(*audits)
        s.update(label="✓ Tono IA evaluado", state="complete")

    # --- PASO 3: SUBTEMAS Y TEMAS ---
    with st.status("Paso 3 · Clasificando Subtemas y Temas...", expanded=True) as s:
        pb = st.progress(0)
        
        # Subtemas
        if "Solo Modelos PKL" in mode:
            subtemas = ["N/A"] * len(df)
        else:
            subtemas = ClasificadorSubtema(bn, al).procesar_lote(
                df["_txt"], pb, df[sc].fillna(''), df[tc].fillna('')
            )

        # Temas
        if epkl:
            # Si se subió PKL de Temas, usar las predicciones directas del modelo
            tp = analizar_temas_con_pkl(df["_txt"].tolist(), epkl)
            if tp:
                temas = tp
            else:
                temas = ["N/A"] * len(df)
        elif "Solo Modelos PKL" in mode:
            temas = ["N/A"] * len(df)
        else:
            temas = consolidar_temas(subtemas, df["_txt"].tolist(), pb, bn)

        df['Subtema'] = subtemas
        df['Tema']    = _unificar_tema_por_subtema(temas, subtemas)
        df = aplicar_consistencia_grupos(df, tc, sc)
        s.update(label="✓ Clasificación completada", state="complete")

    # Escribir las 3 columnas adicionales al final en la hoja openpyxl respetando el formato original
    max_col = ws.max_column
    col_tono    = max_col + 1
    col_tema    = max_col + 2
    col_subtema = max_col + 3
    col_contexto = max_col + 4
    col_coincidencia = max_col + 5
    col_origen = max_col + 6

    # Encabezados en negrita
    font_bold = Font(bold=True)
    ws.cell(row=1, column=col_tono, value="Tono IA").font = font_bold
    ws.cell(row=1, column=col_tema, value="Tema").font = font_bold
    ws.cell(row=1, column=col_subtema, value="Subtema").font = font_bold
    ws.cell(row=1, column=col_contexto, value="Contexto analizado").font = font_bold
    ws.cell(row=1, column=col_coincidencia, value="Coincidencia marca").font = font_bold
    ws.cell(row=1, column=col_origen, value="Origen coincidencia").font = font_bold

    # Asignar valores por fila manteniendo la coincidencia exacta
    for idx, row_data in df.iterrows():
        r = idx + 2
        ws.cell(row=r, column=col_tono, value=str(row_data['Tono IA']))
        ws.cell(row=r, column=col_tema, value=str(row_data['Tema']))
        ws.cell(row=r, column=col_subtema, value=str(row_data['Subtema']))
        ws.cell(row=r, column=col_contexto, value=str(row_data['Contexto analizado']))
        ws.cell(row=r, column=col_coincidencia, value=str(row_data['Coincidencia marca']))
        ws.cell(row=r, column=col_origen, value=str(row_data['Origen coincidencia']))

    buf_out = io.BytesIO()
    wb.save(buf_out)

    ci = (st.session_state['tokens_input']     / 1e6) * PRICE_INPUT_1M
    co = (st.session_state['tokens_output']    / 1e6) * PRICE_OUTPUT_1M
    ce = (st.session_state['tokens_embedding'] / 1e6) * PRICE_EMBEDDING_1M

    cost_str = f"${ci + co + ce:.4f} USD"
    time_str = f"{time.time() - t0:.0f}s"

    return buf_out.getvalue(), df, cost_str, time_str


def render_custom_excel_tab():
    st.markdown('<div class="sec-label">Análisis de Excel Personalizado</div>', unsafe_allow_html=True)
    st.caption("Sube cualquier archivo Excel (.xlsx). Al finalizar se descargarán los mismos datos y formato original con 3 nuevas columnas añadidas al final: **Tono IA**, **Tema** y **Subtema**.")

    if 'custom_result_bytes' in st.session_state:
        st.markdown(
            '<div class="success-banner"><div class="success-icon">✓</div>'
            '<div><div class="success-title">Análisis de Excel Finalizado</div>'
            '<div class="success-sub">Se han añadido las 3 columnas al final del Excel original manteniendo su formato.</div></div></div>',
            unsafe_allow_html=True
        )
        c1, c2 = st.columns(2)
        c1.metric("Costo estimado", st.session_state.get('custom_cost', "$0.00"))
        c2.metric("Tiempo de ejecución", st.session_state.get('custom_time', "0s"))

        if 'custom_df_preview' in st.session_state:
            st.markdown("##### Vista previa del archivo (primeras filas con columnas añadidas):")
            st.dataframe(st.session_state.custom_df_preview.head(10), use_container_width=True)

        st.download_button(
            "⬇ Descargar Excel Actualizado",
            data=st.session_state.custom_result_bytes,
            file_name=f"Analisis_{st.session_state.get('custom_filename', 'Personalizado.xlsx')}",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )

        if st.button("Nuevo análisis personalizado"):
            for k in ('custom_result_bytes', 'custom_df', 'custom_filename', 'custom_cost', 'custom_time', 'custom_df_preview'):
                if k in st.session_state:
                    del st.session_state[k]
            st.rerun()
        return

    if 'custom_df' not in st.session_state:
        f = st.file_uploader("Sube cualquier archivo Excel (.xlsx)", type=["xlsx"], key="custom_uploader")
        if f:
            try:
                bytes_data = f.getvalue()
                df_temp = pd.read_excel(io.BytesIO(bytes_data))
                st.session_state.custom_df       = df_temp
                st.session_state.custom_bytes    = bytes_data
                st.session_state.custom_filename = f.name
                st.rerun()
            except Exception as e:
                st.error(f"Error al leer el archivo Excel: {e}")
    else:
        st.success(f"📁 Archivo cargado: **{st.session_state.custom_filename}** ({len(st.session_state.custom_df)} filas)")

        cols = st.session_state.custom_df.columns.tolist()

        with st.form("custom_form"):
            st.markdown('<div class="sec-label">Selección de Columnas</div>', unsafe_allow_html=True)
            c_col1, c_col2 = st.columns(2)
            tc = c_col1.selectbox("Columna que contiene el TÍTULO", cols, index=_default_text_column_index(cols, ['Título', 'Titulo', 'Titular', 'Headline'], 0))
            sc = c_col2.selectbox("Columna que contiene el RESUMEN / CUERPO", cols, index=_default_text_column_index(cols, ['CuerpoEs', 'Resumen - Aclaración', 'Resumen - Aclaracion', 'Resumen', 'Cuerpo', 'Descripción', 'Descripcion'], 1))

            st.markdown('<div class="sec-label">Configuración del Análisis</div>', unsafe_allow_html=True)
            cl, cr = st.columns([3, 2])
            with cl:
                bn  = st.text_input("Marca principal", placeholder="Ej: Bancolombia", key="custom_bn")
                bat = st.text_input("Alias (separados por ;)", placeholder="Ej: Grupo Bancolombia;Ban", key="custom_ba")
            with cr:
                mode = st.radio(
                    "Modo de análisis",
                    ["API de OpenAI", "Híbrido (PKL + API)", "Solo Modelos PKL"],
                    index=0, key="custom_mode"
                )

            tpkl, epkl = None, None
            st.markdown('<div class="sec-label">Modelos PKL (Opcionales)</div>', unsafe_allow_html=True)
            p1, p2 = st.columns(2)
            tpkl = p1.file_uploader("Modelo Sentimiento / Tono (.pkl)", type=["pkl"], key="custom_tpkl")
            epkl = p2.file_uploader("Modelo Temas (.pkl)", type=["pkl"], key="custom_epkl")

            if st.form_submit_button("▶ Iniciar análisis personalizado", use_container_width=True, type="primary"):
                if not bn.strip():
                    st.error("Ingresa el nombre de la marca principal.")
                elif "Solo Modelos PKL" in mode and not (tpkl or epkl):
                    st.error("Seleccionaste 'Solo Modelos PKL', por favor adjunta al menos un archivo .pkl para continuar.")
                else:
                    if "API" in mode or "Híbrido" in mode:
                        try:
                            openai.api_key = st.secrets["OPENAI_API_KEY"]
                            openai.aiosession.set(None)
                        except:
                            st.error("OPENAI_API_KEY no encontrada en st.secrets.")
                            st.stop()

                    al = [a.strip() for a in bat.split(";") if a.strip()]

                    with st.spinner("Procesando Excel personalizado..."):
                        res_bytes, res_df, cost_str, time_str = asyncio.run(
                            run_custom_excel_async(
                                st.session_state.custom_bytes,
                                tc, sc, bn, al,
                                mode=mode, tpkl=tpkl, epkl=epkl
                            )
                        )

                        st.session_state.custom_result_bytes = res_bytes
                        st.session_state.custom_df_preview   = res_df
                        st.session_state.custom_cost         = cost_str
                        st.session_state.custom_time         = time_str
                        st.rerun()

        if st.button("Subir otro archivo Excel"):
            for k in ('custom_df', 'custom_bytes', 'custom_filename', 'custom_result_bytes', 'custom_cost', 'custom_time', 'custom_df_preview'):
                if k in st.session_state:
                    del st.session_state[k]
            st.rerun()


# ======================================
# Main
# ======================================
async def run_sentiment_only_async(df, title_col, summary_col, brand, aliases, pkl_file=None):
    details = [extraer_contexto_marca_detallado(r.get(title_col, ''), r.get(summary_col, ''), brand, aliases) for _, r in df.iterrows()]
    df = df.copy()
    idx = [i for i, d in enumerate(details) if d['contexto']]
    results = [{'tono':'Neutro','confianza':'Alta','justificacion':'La marca no aparece en el título ni en el resumen.'} for _ in details]
    if idx:
        pb = st.progress(0)
        if pkl_file:
            raw = analizar_tono_con_pkl([details[i]['contexto'] for i in idx], pkl_file)
        else:
            raw = await ClasificadorTono(brand, aliases).procesar_lote_async(pd.Series([details[i]['contexto'] for i in idx]), pb, pd.Series([df.iloc[i][summary_col] for i in idx]), pd.Series([df.iloc[i][title_col] for i in idx]))
        for i, r in zip(idx, raw or []): results[i].update(r)
    df['Tono IA'] = [r.get('tono','Neutro') for r in results]
    df['Confianza Tono'] = [r.get('confianza','Media') for r in results]
    df['Marca encontrada'] = [d['marca_encontrada'] for d in details]
    df['Contexto analizado'] = [d['contexto'] for d in details]
    df['Coincidencia marca'] = [d['coincidencia'] for d in details]
    df['Origen coincidencia'] = [d['origen'] for d in details]
    return df

def render_sentiment_tab():
    st.markdown('<div class="sec-label">Sentimiento por Marca</div>', unsafe_allow_html=True)
    st.caption('Analiza exclusivamente el impacto reputacional de la marca encontrada en título y resumen.')
    f = st.file_uploader('Sube un Excel (.xlsx)', type=['xlsx'], key='sentiment_uploader')
    if not f: return
    try:
        df = pd.read_excel(io.BytesIO(f.getvalue())); cols = df.columns.tolist()
        with st.form('sentiment_form'):
            tc = st.selectbox('Columna de título', cols, _default_text_column_index(cols, ['Título', 'Titulo', 'Titular', 'Headline'], 0))
            sc = st.selectbox('Columna de resumen / aclaración', cols, _default_text_column_index(cols, ['CuerpoEs', 'Resumen - Aclaración', 'Resumen - Aclaracion', 'Resumen', 'Cuerpo', 'Descripción', 'Descripcion'], 1))
            brand = st.text_input('Marca principal'); alias_text = st.text_input('Alias separados por ;')
            pkl = st.file_uploader('Modelo PKL opcional', type=['pkl'], key='sentiment_pkl')
            submit = st.form_submit_button('Analizar sentimiento', type='primary', use_container_width=True)
        if submit:
            if not brand.strip(): st.error('Ingresa la marca principal.')
            else:
                if not pkl: openai.api_key = st.secrets['OPENAI_API_KEY']
                aliases = [a.strip() for a in alias_text.split(';') if a.strip()]
                with st.spinner('Analizando menciones de la marca...'):
                    result = asyncio.run(run_sentiment_only_async(df, tc, sc, brand.strip(), aliases, pkl))
                st.dataframe(result.head(20), use_container_width=True)
                out = io.BytesIO()
                with pd.ExcelWriter(out, engine='openpyxl') as w: result.to_excel(w, index=False, sheet_name='Sentimiento')
                output_name = f"sentimiento_{_safe_filename_part(brand)}.xlsx"
                st.download_button('Descargar Excel de Sentimiento', out.getvalue(), output_name, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', use_container_width=True, type='primary')
    except Exception as e: st.error(f'Error durante el análisis: {e}')

def main():
    load_custom_css()
    if not check_password(): return

    st.markdown("""
    <div class="app-header">
        <div class="app-header-icon">◈</div>
        <div class="app-header-text">
            <div class="app-header-title">Análisis de Sentimiento por Marca</div>
            <div class="app-header-version">Motor reputacional · OpenAI o modelo PKL</div>
        </div>
        <div class="app-header-badge">SENTIMIENTO</div>
    </div>""", unsafe_allow_html=True)

    render_sentiment_tab()

    st.markdown(
        '<div class="footer">Análisis reputacional de sentimiento · Johnathan Cortés ©</div>',
        unsafe_allow_html=True
    )

if __name__ == "__main__":
    main()
